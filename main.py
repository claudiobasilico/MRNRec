"""
SIRIO - Sistema di Riconciliazione — v1.0
Riconciliazione intelligente fra fatture e dichiarazioni doganali
Singolo modulo FastAPI — porta 8002
"""

# ════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ════════════════════════════════════════════════════════════════════════════
import sys, os, json, re, unicodedata, logging, hashlib, secrets, csv
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any
from io import BytesIO, TextIOWrapper

import uvicorn
from fastapi import FastAPI, UploadFile, File, Form, Request, HTTPException, Depends
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import xlrd


# ════════════════════════════════════════════════════════════════════════════
# BASE DIR (funziona da .py e da PyInstaller)
# ════════════════════════════════════════════════════════════════════════════
def _base() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent

BASE_DIR = _base()
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

TOKENS_FILE  = DATA_DIR / "tokens.json"
LOGS_FILE    = DATA_DIR / "activity_log.jsonl"
USAGE_FILE   = DATA_DIR / "usage.json"
PROFILES_FILE = DATA_DIR / "profiles.json"


# ════════════════════════════════════════════════════════════════════════════
# LOGGING
# ════════════════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(DATA_DIR / "server.log", encoding="utf-8")
    ]
)
log = logging.getLogger("mrn_rec")


# ════════════════════════════════════════════════════════════════════════════
# TOKEN STORE — struttura: { token: { "name": str, "max_rows": int, "active": bool } }
# USAGE — conta le righe di fatture elaborate: { token: righe_elaborate }
# ════════════════════════════════════════════════════════════════════════════
def load_tokens() -> dict:
    if TOKENS_FILE.exists():
        with open(TOKENS_FILE, encoding="utf-8") as f:
            return json.load(f)
    # Primo avvio: crea token admin
    admin_token = secrets.token_urlsafe(24)
    initial = {
        admin_token: {
            "name": "SIRIO_ADMIN",
            "max_rows": 1000,       # 1000 righe di default
            "active": True,
            "created_at": datetime.now().isoformat(),
            "notes": "Token amministratore"
        }
    }
    _save_tokens(initial)
    log.warning(f"⚠ PRIMO AVVIO — Token admin generato: {admin_token}")
    log.warning(f"⚠ Salvato in: {TOKENS_FILE}")
    return initial

def _save_tokens(tokens: dict):
    with open(TOKENS_FILE, "w", encoding="utf-8") as f:
        json.dump(tokens, f, indent=2, ensure_ascii=False)

TOKENS: dict = load_tokens()


# ════════════════════════════════════════════════════════════════════════════
# USAGE STORE — { token: int }  contatore utilizzi
# ════════════════════════════════════════════════════════════════════════════
def load_usage() -> dict:
    if USAGE_FILE.exists():
        with open(USAGE_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_usage(usage: dict):
    with open(USAGE_FILE, "w", encoding="utf-8") as f:
        json.dump(usage, f, indent=2)

USAGE: dict = load_usage()


# ════════════════════════════════════════════════════════════════════════════
# ACTIVITY LOG — JSONL append
# ════════════════════════════════════════════════════════════════════════════
def write_log(token: str, action: str, details: dict):
    entry = {
        "ts":      datetime.now().isoformat(timespec="seconds"),
        "token":   token,
        "user":    TOKENS.get(token, {}).get("name", "unknown"),
        "action":  action,
        **details
    }
    with open(LOGS_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


# ════════════════════════════════════════════════════════════════════════════
# AUTH DEPENDENCY
# ════════════════════════════════════════════════════════════════════════════
def require_token(request: Request) -> str:
    token = (
        request.headers.get("X-Token") or
        request.query_params.get("token")
    )
    if not token or token not in TOKENS:
        raise HTTPException(status_code=401, detail="Token non valido o mancante")
    
    info = TOKENS[token]
    if not info.get("active", True):
        raise HTTPException(status_code=403, detail="Token disabilitato")

    max_rows = info.get("max_rows", 0)
    if max_rows > 0:
        used_rows = USAGE.get(token, 0)
        if used_rows >= max_rows:
            raise HTTPException(
                status_code=429,
                detail=f"Limite righe raggiunto ({used_rows}/{max_rows}). Contatta SIRIO per estendere il token."
            )
    return token

def consume_rows(token: str, num_rows: int):
    """Incrementa contatore righe elaborate solo se token ha limite."""
    if TOKENS[token].get("max_rows", 0) > 0:
        USAGE[token] = USAGE.get(token, 0) + num_rows
        save_usage(USAGE)


# ════════════════════════════════════════════════════════════════════════════
# NORMALIZZAZIONE
# ════════════════════════════════════════════════════════════════════════════
_PREFIX_RE = re.compile(
    r"^(FT|INV|FAT|FATT|NOTE?|NC|ND|PRF|PRO|DDT|OR?D|REF|N)\s*[.\-/]?\s*",
    re.IGNORECASE
)
_NONALNUM = re.compile(r"[^A-Z0-9]")

def normalize_invoice(raw: str) -> dict:
    """Produce varianti di normalizzazione di un numero fattura."""
    if not raw:
        return {}
    s = str(raw).strip().upper()
    # rimuove diacritici
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    raw_norm = s
    # rimuove prefissi
    s_clean = _PREFIX_RE.sub("", s)
    # solo alfanumerico
    only_alnum = _NONALNUM.sub("", s_clean)
    # solo numerico
    digits = re.sub(r"[^0-9]", "", only_alnum)
    # anno embedded (es. 2024/001 → anno=2024, num=001)
    year_match = re.search(r"(20\d{2})", digits)
    year = year_match.group(1) if year_match else None
    num_core = re.sub(r"20\d{2}", "", digits).lstrip("0") if year else digits.lstrip("0")
    # suffix finale >= 5 cifre
    suffix = digits[-5:] if len(digits) >= 5 else digits
    return {
        "raw":          raw_norm,
        "clean":        s_clean,
        "alnum":        only_alnum,
        "numeric_core": num_core,
        "suffix":       suffix,
        "year":         year,
        "digits":       digits,
    }

def parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val

    # Gestisci numeri Excel (OLE Automation Date)
    if isinstance(val, (int, float)):
        try:
            # Excel epoch è 1899-12-30
            # 46085 corrisponde a 2026-03-04
            excel_epoch = datetime(1899, 12, 30)
            date_obj = excel_epoch + timedelta(days=int(val))
            return date_obj.date()
        except:
            pass

    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def parse_amount(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ".").replace(" ", ""))
    except:
        return None


# ════════════════════════════════════════════════════════════════════════════
# EXCEL READER — struttura variabile, header auto-detection
# ════════════════════════════════════════════════════════════════════════════
_HEADER_ALIASES = {
    # MRN doganale
    "mrn":           ["mrn", "numero mrn", "dichiarazione", "num_dichiarazione", "n. dichiarazione"],
    "fattura_dog":   ["fattura", "rif. fattura", "riferimento fattura", "invoice", "num. fattura",
                      "numero fattura", "ref fattura", "fattura cliente", "n. fattura"],
    "data_dog":      ["data accettazione", "data uscita effettiva", "data uscita", "data sdoganamento",
                      "data dichiarazione", "data", "data mrn", "date"],
    "importo_dog":   ["importo", "valore", "importo merce", "valore fattura", "amount",
                      "imponibile", "valore merce"],
    "cliente_dog":   ["destinatario", "consegnatario", "cliente", "intestatario", "customer", "shipper",
                      "mittente", "ragione sociale"],
    "paese_dog":     ["paese", "country", "paese origine", "paese provenienza", "nazione"],
    "valuta_dog":    ["valuta", "currency", "divisa"],

    # Fatture cliente
    "num_fattura":   ["numero fattura", "n. fattura", "n fattura", "fattura n", "fattura nr",
                      "invoice number", "invoice no", "num fattura", "rif fattura", "numero"],
    "data_fattura":  ["data fattura", "data emissione", "invoice date", "data", "emessa il"],
    "importo_fat":   ["importo", "totale", "importo fattura", "totale fattura", "amount",
                      "valore", "imponibile", "totale imponibile"],
    "cliente_fat":   ["cliente", "intestatario", "ragione sociale cliente", "customer",
                      "destinatario", "nome cliente"],
    "paese_fat":     ["paese", "country", "nazione", "stato"],
    "valuta_fat":    ["valuta", "currency", "divisa"],
}

def _match_header(col_name: str, field_candidates: list) -> bool:
    s = str(col_name).strip().lower()
    for c in field_candidates:
        if c in s or s in c:
            return True
    return False

def detect_columns(headers: list, aliases: dict) -> dict:
    """Ritorna {field: col_name} per i campi riconosciuti."""
    mapping = {}
    for field, candidates in aliases.items():
        for h in headers:
            if _match_header(h, candidates):
                if field not in mapping:
                    mapping[field] = h
    return mapping

def load_profiles() -> dict:
    """Carica TUTTI i profili (struttura: token → profili)."""
    if PROFILES_FILE.exists():
        try:
            with open(PROFILES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            log.warning(f"Errore lettura profili: {str(e)}")
    return {}

def get_user_profiles(token: str) -> dict:
    """Ritorna i profili dell'utente specifico (basato su token)."""
    all_profiles = load_profiles()
    return all_profiles.get(token, {})

def save_profile(token: str, name: str, col_mapping: dict) -> bool:
    """Salva un profilo per un utente specifico (basato su token)."""
    try:
        profiles = load_profiles()
        if token not in profiles:
            profiles[token] = {}
        profiles[token][name] = col_mapping
        with open(PROFILES_FILE, 'w', encoding='utf-8') as f:
            json.dump(profiles, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        log.error(f"Errore salvataggio profilo '{name}' per token {token}: {str(e)}")
        return False

def get_profile(token: str, name: str) -> Optional[dict]:
    """Legge un profilo specifico per token e nome."""
    user_profiles = get_user_profiles(token)
    return user_profiles.get(name)

def delete_profile(token: str, name: str) -> bool:
    """Elimina un profilo specifico."""
    try:
        profiles = load_profiles()
        if token in profiles and name in profiles[token]:
            del profiles[token][name]
            # Se l'utente non ha più profili, rimuovi la chiave token
            if not profiles[token]:
                del profiles[token]
            with open(PROFILES_FILE, 'w', encoding='utf-8') as f:
                json.dump(profiles, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        log.error(f"Errore eliminazione profilo '{name}': {str(e)}")
        return False

def read_excel_rows(data: bytes, filename: str = "") -> list[dict]:
    """Legge Excel (.xls, .xlsx) o CSV, auto-detecta header, ritorna lista di dict."""
    ext = filename.lower().split('.')[-1] if filename else ""

    # CSV
    if ext == "csv":
        try:
            reader = csv.DictReader(TextIOWrapper(BytesIO(data), encoding='utf-8'))
            rows = []
            for row in reader:
                # Rimuovi chiavi vuote
                row = {k.strip(): v for k, v in row.items() if k.strip()}
                rows.append(row)
            return rows
        except Exception as e:
            raise HTTPException(400, f"Errore lettura CSV: {str(e)}")

    # .xls (formato legacy binario)
    if ext == "xls":
        try:
            wb = xlrd.open_workbook(file_contents=data)
            ws = wb.sheet_by_index(0)
            rows = []
            for row_idx in range(ws.nrows):
                rows.append(ws.row_values(row_idx))
        except Exception as e:
            raise HTTPException(400, f"File .xls non valido: {str(e)}")
    else:
        # .xlsx (formato moderno)
        try:
            wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            raise HTTPException(400, f"File Excel non valido: {str(e)}")

    if len(rows) < 2:
        return []

    # Trova riga header (prima riga non completamente vuota)
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        if any(c is not None for c in row):
            header_row = [str(c).strip() if c is not None else "" for c in row]
            header_idx = i
            break
    if header_row is None:
        return []

    result = []
    for row in rows[header_idx + 1:]:
        if all(c is None for c in row):
            continue
        # Assicura che tutte le colonne dell'header siano sempre presenti
        d = {}
        for i, h in enumerate(header_row):
            d[h] = row[i] if i < len(row) else None
        result.append(d)
    return result


# ════════════════════════════════════════════════════════════════════════════
# MATCHING ENGINE
# ════════════════════════════════════════════════════════════════════════════
MATCH_CLASSES = [
    (95, "CERTO"),
    (85, "PROBABILE"),
    (70, "POSSIBILE"),
    (50, "DEBOLE"),
    (0,  "NESSUN MATCH"),
]

def classify(score: float) -> str:
    for threshold, label in MATCH_CLASSES:
        if score >= threshold:
            return label
    return "NESSUN MATCH"

def _levenshtein(a: str, b: str) -> int:
    """Distanza di edit semplice."""
    if not a: return len(b)
    if not b: return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            curr.append(min(prev[j] + 1, curr[-1] + 1, prev[j-1] + (ca != cb)))
        prev = curr
    return prev[-1]

def normalize_company_name(raw: str) -> str:
    """
    Normalizza il nome aziendale per matching fuzzy:
    - Rimuove diacritici (ä→a, ö→o, é→e, etc.)
    - Normalizza abbreviazioni comuni (GmbH, Ltd, Inc, etc.)
    - Lowercaser e pulizia spazi
    - Rimuove i suffissi legali comuni
    """
    s = str(raw).upper().strip() if raw else ""

    # Normalizza diacritici unicode
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')

    # Mappa abbreviazioni comuni a forme normalizzate
    abbr_map = {
        'GMBH': 'GESELLSCHAFT',
        'GMBH & CO. KG': 'GESELLSCHAFT',
        'AG': 'AKTIENGESELLSCHAFT',
        'SA': 'SOCIETE ANONYME',
        'SARL': 'SARL',
        'EIRL': 'EIRL',
        'SCRL': 'SCRL',
        'CVBA': 'CVBA',
        'LTD': 'LIMITED',
        'LLC': 'COMPANY',
        'INC': 'INCORPORATED',
        'CORP': 'CORPORATION',
        'CORP.': 'CORPORATION',
        'PLC': 'PUBLIC',
        'NV': 'NAAMLOZE VENNOOTSCHAP',
        'BV': 'BESLOTEN VENNOOTSCHAP',
        'KG': 'KOMMANDITGESELLSCHAFT',
        'OHG': 'OFFENE HANDELSGESELLSCHAFT',
        'EV': 'EINGETRAGENER VEREIN',
        'EURL': 'ENTREPRISE UNIPERSONNELLE',
        'SAS': 'SOCIETE',
        'SASU': 'SOCIETE',
        'SELARL': 'SELARL',
        'SPRL': 'SPRL',
    }

    for abbr, expanded in abbr_map.items():
        s = re.sub(r'\b' + re.escape(abbr) + r'\b', expanded, s)

    # Pulisci: rimuovi punteggiatura, normalizza spazi
    s = re.sub(r'[.,;:&\-/\(\)\[\]{}]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()

    # Rimuovi articoli comuni (non viene fatto altrimenti)
    for article in ['THE ', 'DIE ', 'DER ', 'DAS ', 'LA ', 'LE ', 'IL ', 'LO ']:
        s = s.replace(article, '')

    return s

def match_company_names(fat_client: str, mrn_client: str) -> tuple[float, str]:
    """
    Match intelligente tra ragioni sociali con fuzzy matching.
    Ritorna (score, metodo).
    Score 0-10.
    """
    if not fat_client or not mrn_client:
        return 0.0, "no_company_names"

    fat_client = fat_client.upper().strip()
    mrn_client = mrn_client.upper().strip()

    # 1. Match esatto
    if fat_client == mrn_client:
        return 10.0, "company_exact"

    # 2. Normalizza e prova match esatto
    fat_norm = normalize_company_name(fat_client)
    mrn_norm = normalize_company_name(mrn_client)

    if fat_norm == mrn_norm:
        return 9.5, "company_normalized"

    # 3. Tokenize e overlap con normali
    fat_tokens = set(re.split(r"\W+", fat_norm)) - {"SRL", "SPA", "SAS", "SOCIETE", "GESELLSCHAFT", "COMPANY", "LIMITED", "INCORPORATED", "CORPORATION", ""}
    mrn_tokens = set(re.split(r"\W+", mrn_norm)) - {"SRL", "SPA", "SAS", "SOCIETE", "GESELLSCHAFT", "COMPANY", "LIMITED", "INCORPORATED", "CORPORATION", ""}

    if not fat_tokens or not mrn_tokens:
        return 0.0, "no_company_tokens"

    overlap = fat_tokens & mrn_tokens
    if overlap:
        # Score proporzionale all'overlap
        overlap_score = len(overlap) / max(len(fat_tokens), len(mrn_tokens))
        c_score = overlap_score * 8.0  # Max 8 punti per token overlap
        return min(8.0, c_score), f"company_tokens_overlap"

    # 4. Fuzzy match: Levenshtein su token principali
    # Prova match tra tutti i token (non solo il più lungo)
    best_ratio = 0.0
    for fat_tok in fat_tokens:
        for mrn_tok in mrn_tokens:
            if len(fat_tok) >= 4 and len(mrn_tok) >= 4:
                dist = _levenshtein(fat_tok, mrn_tok)
                ratio = 1.0 - (dist / max(len(fat_tok), len(mrn_tok)))
                best_ratio = max(best_ratio, ratio)

    # Se nessun token abbastanza lungo, prova su interi nomi normalizzati
    if best_ratio == 0.0 and fat_norm and mrn_norm:
        dist = _levenshtein(fat_norm, mrn_norm)
        best_ratio = 1.0 - (dist / max(len(fat_norm), len(mrn_norm)))

    if best_ratio >= 0.85:  # 85% di somiglianza
        c_score = best_ratio * 7.0
        return min(7.0, c_score), "company_fuzzy_high"
    elif best_ratio >= 0.75:  # 75% di somiglianza
        c_score = best_ratio * 5.0
        return min(5.0, c_score), "company_fuzzy_mid"
    elif best_ratio >= 0.65:  # 65% di somiglianza
        c_score = best_ratio * 3.0
        return min(3.0, c_score), "company_fuzzy_low"

    return 0.0, "company_no_match"

def _invoice_score(fat_norm: dict, dog_ref_raw: str) -> tuple[float, str]:
    """
    Calcola il punteggio di match fattura (max 60) e il metodo.
    Ritorna (score, method_description).
    """
    if not dog_ref_raw or not fat_norm:
        return 0.0, "nessun_riferimento"

    # Tokenizza il campo doganale (può contenere più fatture separate da ; , / |)
    tokens_raw = re.split(r"[;,/|\s]+", str(dog_ref_raw).upper())
    tokens = [t.strip() for t in tokens_raw if t.strip()]

    best_score = 0.0
    best_method = "no_match"

    fat_raw  = fat_norm.get("raw", "")
    fat_aln  = fat_norm.get("alnum", "")
    fat_core = fat_norm.get("numeric_core", "")
    fat_suf  = fat_norm.get("suffix", "")
    fat_year = fat_norm.get("year")

    for tok in tokens:
        tok_norm = normalize_invoice(tok)
        tok_aln  = tok_norm.get("alnum", "")
        tok_core = tok_norm.get("numeric_core", "")
        tok_suf  = tok_norm.get("suffix", "")
        tok_year = tok_norm.get("year")

        # 1. Match esatto raw
        if fat_raw == tok.strip():
            return 60.0, "match_esatto_raw"

        # 2. Match esatto normalizzato (alnum)
        if fat_aln and tok_aln and fat_aln == tok_aln:
            return 58.0, "match_esatto_normalizzato"

        # 3. Match numero + anno
        if fat_core and tok_core and fat_core == tok_core:
            if fat_year and tok_year and fat_year == tok_year:
                s, m = 56.0, "match_numero_anno"
            elif fat_year or tok_year:
                s, m = 40.0, "match_numero_anno_parziale"
            else:
                s, m = 50.0, "match_numero_core"
            if s > best_score:
                best_score, best_method = s, m

        # 4. Match suffix finale (>= 5 cifre)
        if fat_suf and tok_suf and len(fat_suf) >= 5 and fat_suf == tok_suf:
            s, m = 45.0, "match_suffix_5cifre"
            if s > best_score:
                best_score, best_method = s, m

        # 5. Match suffix: numero dogana è gli ultimi N digit del numero fattura
        # Es: fattura 289, dogana 89 → ultimi 2 digit matchano
        if fat_core and tok_core and len(tok_core) >= 2 and len(fat_core) >= len(tok_core):
            if fat_core.endswith(tok_core):
                s, m = 35.0, "match_suffix_numero"
                if s > best_score:
                    best_score, best_method = s, m

        # 6. Match con errore umano (Levenshtein <= 2 su alnum >= 6 char)
        if fat_aln and tok_aln and len(fat_aln) >= 6 and len(tok_aln) >= 6:
            dist = _levenshtein(fat_aln, tok_aln)
            if dist == 1:
                s, m = 42.0, "match_levenshtein_1"
                if s > best_score:
                    best_score, best_method = s, m
            elif dist == 2:
                s, m = 30.0, "match_levenshtein_2"
                if s > best_score:
                    best_score, best_method = s, m

    return best_score, best_method


def score_pair(fat: dict, mrn: dict, fat_col: dict, dog_col: dict) -> dict:
    """
    Calcola score completo tra una fattura e un MRN.
    Ritorna: { score, class, method, explanation, mrn_id, score_inv, score_data, score_amt, score_client, score_paese }
    """
    total = 0.0
    inv_score = 0.0      # Numero fattura (max 60)
    data_score = 0.0     # Data (max 15)
    amt_score = 0.0      # Importo (max 10)
    client_score = 0.0   # Cliente (max 10)
    paese_score = 0.0    # Paese (max 5)
    notes = []
    methods = []

    # ── Numero fattura (max 60) ─────────────────────────────────────────────
    fat_num_val = ""
    dog_ref_val = ""
    if "num_fattura" in fat_col:
        fat_num_val = str(fat.get(fat_col["num_fattura"], ""))
    if "fattura_dog" in dog_col:
        dog_ref_val = str(mrn.get(dog_col["fattura_dog"], ""))

    fat_norm = normalize_invoice(fat_num_val)
    inv_score, inv_method = _invoice_score(fat_norm, dog_ref_val)
    total += inv_score
    methods.append(inv_method)
    if inv_score > 0:
        notes.append(f"Fattura: '{fat_num_val}' ↔ rif.dogana: '{dog_ref_val}' [{inv_method}, +{inv_score:.0f}]")

    # ── Data (max 15) ─────────────────────────────────────────────────────────
    fat_date = None
    mrn_date = None
    if "data_fattura" in fat_col:
        fat_date = parse_date(fat.get(fat_col["data_fattura"]))
    if "data_dog" in dog_col:
        mrn_date = parse_date(mrn.get(dog_col["data_dog"]))

    if fat_date and mrn_date:
        # Vincolo: data MRN >= data fattura
        if mrn_date < fat_date:
            notes.append(f"⚠ Data MRN ({mrn_date}) < Data fattura ({fat_date}) — penalità -10")
            total -= 10
            methods.append("penalità_data_antecedente")
        else:
            delta = (mrn_date - fat_date).days
            # Max score within 5 days, min score within 90 days
            if delta <= 5:
                d_score = 15.0
            elif delta <= 20:
                d_score = 12.0
            elif delta <= 40:
                d_score = 9.0
            elif delta <= 60:
                d_score = 6.0
            elif delta <= 90:
                d_score = 2.0
            else:
                d_score = 0.0
            if d_score > 0:
                total += d_score
                data_score = d_score
                notes.append(f"Data: Δ{delta}gg +{d_score:.0f}")
                methods.append("data")

    # ── Cliente (max 10) ──────────────────────────────────────────────────────
    fat_client = ""
    mrn_client = ""
    if "cliente_fat" in fat_col:
        fat_client = str(fat.get(fat_col["cliente_fat"], "")).upper().strip()
    if "cliente_dog" in dog_col:
        mrn_client = str(mrn.get(dog_col["cliente_dog"], "")).upper().strip()

    if fat_client and mrn_client:
        c_score, c_method = match_company_names(fat_client, mrn_client)
        if c_score > 0:
            total += c_score
            client_score = c_score
            notes.append(f"Cliente: '{fat_client}' ↔ '{mrn_client}' [{c_method}, +{c_score:.1f}]")
            methods.append("cliente")

    # ── Paese (max 5) ─────────────────────────────────────────────────────────
    fat_paese = ""
    mrn_paese = ""
    if "paese_fat" in fat_col:
        fat_paese = str(fat.get(fat_col["paese_fat"], "")).upper().strip()[:2]
    if "paese_dog" in dog_col:
        mrn_paese = str(mrn.get(dog_col["paese_dog"], "")).upper().strip()[:2]

    if fat_paese and mrn_paese and fat_paese == mrn_paese:
        paese_score = 5.0
        total += paese_score
        notes.append(f"Paese: {fat_paese} +5")
        methods.append("paese")

    # ── Importo (max 10) ──────────────────────────────────────────────────────
    fat_amt = None
    mrn_amt = None
    if "importo_fat" in fat_col:
        fat_amt = parse_amount(fat.get(fat_col["importo_fat"]))
    if "importo_dog" in dog_col:
        mrn_amt = parse_amount(mrn.get(dog_col["importo_dog"]))

    if fat_amt and mrn_amt and fat_amt > 0:
        # Dogana ha valore statistico, non importo fattura
        # Max score when equal, min score when differs by ±1000 euro
        diff = abs(fat_amt - mrn_amt)
        if diff == 0:
            a_score = 10.0
        elif diff <= 500:
            a_score = 8.0
        elif diff <= 1000:
            a_score = 4.0
        else:
            a_score = 0.0
        if a_score > 0:
            total += a_score
            amt_score = a_score
            notes.append(f"Importo: {fat_amt:.2f} ↔ {mrn_amt:.2f} (Δ{diff:.2f}€) +{a_score:.0f}")
            methods.append("importo")

    # Clamp 0-100
    total = max(0.0, min(100.0, total))

    mrn_id = ""
    if "mrn" in dog_col:
        mrn_id = str(mrn.get(dog_col["mrn"], ""))

    # Debug: log final score for invoice 289
    if fat_num_val and "289" in str(fat_num_val):
        log.warning(f"[score_pair FINAL] FATTURA 289 → MRN {mrn_id}: SCORE={total:.1f} ({classify(total)}) | {' | '.join(notes)}")

    return {
        "mrn":         mrn_id,
        "score_inv":   round(inv_score, 1),    # Numero fattura
        "score_data":  round(data_score, 1),   # Data
        "score_amt":   round(amt_score, 1),    # Importo
        "score_client": round(client_score, 1), # Cliente
        "score_paese": round(paese_score, 1),  # Paese
        "score":       round(total, 1),        # Totale
        "class":       classify(total),
        "explanation": "; ".join(notes) if notes else "Nessun elemento corrispondente",
        "method":      " | ".join(m for m in methods if m not in ("no_match",)),
        "fat_date":    str(fat_date) if fat_date else "",
        "mrn_date":    str(mrn_date) if mrn_date else "",
        "fat_client":  fat_client,
        "mrn_client":  mrn_client,
        "fat_amount":  fat_amt,
        "mrn_amount":  mrn_amt,
    }


def reconcile(fat_rows: list, mrn_rows: list,
              fat_headers: list, dog_headers: list, col_map: dict = None) -> list:
    """
    Esegue la riconciliazione completa.
    Ritorna lista di fatture, ciascuna con lista di MRN candidati ordinati per score.
    Se col_map è fornito, lo usa al posto dell'auto-detect.
    """
    if col_map:
        fat_col = col_map.get("fat", {})
        dog_col = col_map.get("dog", {})
    else:
        fat_col = detect_columns(fat_headers, {k: v for k, v in _HEADER_ALIASES.items()
                                               if not k.endswith("_dog")})
        dog_col = detect_columns(dog_headers, {k: v for k, v in _HEADER_ALIASES.items()
                                               if k.endswith("_dog") or k == "mrn"})

    results = []
    for fat in fat_rows:
        candidates = []
        for mrn in mrn_rows:
            match = score_pair(fat, mrn, fat_col, dog_col)
            if match["score"] >= 50:
                candidates.append(match)

        candidates.sort(key=lambda x: x["score"], reverse=True)

        # Recupera numero e data fattura per display
        fat_num_display = fat.get(fat_col.get("num_fattura"), "") if "num_fattura" in fat_col else ""
        fat_date_display = fat.get(fat_col.get("data_fattura"), "") if "data_fattura" in fat_col else ""
        fat_client_display = fat.get(fat_col.get("cliente_fat"), "") if "cliente_fat" in fat_col else ""
        fat_amount_display = fat.get(fat_col.get("importo_fat"), "") if "importo_fat" in fat_col else ""
        fat_num_display = str(fat_num_display) if fat_num_display else ""
        fat_date_display = str(fat_date_display) if fat_date_display else ""
        fat_client_display = str(fat_client_display) if fat_client_display else ""
        fat_amount_display = str(fat_amount_display) if fat_amount_display else ""

        results.append({
            "fattura_num":    fat_num_display,
            "fattura_data":   fat_date_display,
            "fattura_cliente":fat_client_display,
            "fattura_importo":fat_amount_display,
            "candidati":      candidates,
            "raw":            fat,
        })
    return results


# ════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ════════════════════════════════════════════════════════════════════════════
COLOR_MAP = {
    "CERTO":       "00C851",
    "PROBABILE":   "33B5E5",
    "POSSIBILE":   "FFBB33",
    "DEBOLE":      "FF4444",
    "NESSUN MATCH":"AAAAAA",
}

def _safe_str(val) -> str:
    """Converte qualsiasi valore a stringa in modo sicuro."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, (date, datetime)):
        return val.isoformat()
    return str(val)

def format_date_for_excel(date_val) -> str:
    """Converte qualsiasi formato di data a DD/MM/YYYY."""
    if not date_val:
        return ""

    try:
        # Se è una stringa che contiene un numero (numero Excel come stringa)
        if isinstance(date_val, str):
            # Prova prima a parsare come numero Excel
            try:
                date_num = float(date_val)
                parsed_date = parse_date(date_num)
                if parsed_date:
                    return parsed_date.strftime("%d/%m/%Y")
            except:
                # Altrimenti è una stringa ISO o altro formato
                pass

            # Prova a parsare direttamente come stringa
            parsed_date = parse_date(date_val)
            if parsed_date:
                return parsed_date.strftime("%d/%m/%Y")

        # Se è un numero (float)
        elif isinstance(date_val, (int, float)):
            parsed_date = parse_date(date_val)
            if parsed_date:
                return parsed_date.strftime("%d/%m/%Y")

        return str(date_val)
    except Exception as e:
        return str(date_val)


def build_output_excel(results: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Riconciliazione MRN"

    headers = [
        "Fattura N°", "Data Fattura", "Cliente Fattura", "Importo Fattura",
        "MRN", "Data MRN", "Cliente Dogana", "Importo Dogana",
        "Score Fattura", "Score Data", "Score Importo", "Score Cliente", "Score Paese",
        "Score Totale", "Classe", "Spiegazione"
    ]
    bold = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = hdr_fill
        cell.alignment = center

    ws.row_dimensions[1].height = 28

    row_num = 2
    for fat in results:
        # Converti data fattura a formato uniforme DD/MM/YYYY
        fat_date_str = format_date_for_excel(fat.get("fattura_data"))

        if not fat["candidati"]:
            row = [
                _safe_str(fat["fattura_num"]),
                fat_date_str,
                _safe_str(fat["fattura_cliente"]),
                _safe_str(fat["fattura_importo"]),
                "—", "", "", "",
                0, 0, 0, 0, 0,
                0, "NESSUN MATCH", "Nessun MRN candidato"
            ]
            ws.append(row)
            fill = PatternFill("solid", fgColor="DDDDDD")
            for cell in ws[row_num]:
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
            row_num += 1
        else:
            for i, cand in enumerate(fat["candidati"]):
                mrn_date_str = format_date_for_excel(cand.get("mrn_date"))
                row = [
                    _safe_str(fat["fattura_num"]) if i == 0 else "",
                    fat_date_str if i == 0 else "",
                    _safe_str(fat["fattura_cliente"]) if i == 0 else "",
                    _safe_str(fat["fattura_importo"]) if i == 0 else "",
                    _safe_str(cand["mrn"]),
                    mrn_date_str,
                    _safe_str(cand["mrn_client"]),
                    _safe_str(cand["mrn_amount"]),
                    _safe_str(cand.get("score_inv", 0)),
                    _safe_str(cand.get("score_data", 0)),
                    _safe_str(cand.get("score_amt", 0)),
                    _safe_str(cand.get("score_client", 0)),
                    _safe_str(cand.get("score_paese", 0)),
                    _safe_str(cand["score"]),
                    _safe_str(cand["class"]),
                    _safe_str(cand["explanation"]),
                ]
                ws.append(row)
                fgColor = COLOR_MAP.get(cand["class"], "FFFFFF")
                score_fill = PatternFill("solid", fgColor=fgColor)
                # Colora le celle Score Totale e Classe
                ws.cell(row_num, 14).fill = score_fill  # Score Totale
                ws.cell(row_num, 15).fill = score_fill  # Classe
                for cell in ws[row_num]:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                row_num += 1

    # Larghezze colonne
    # Ordine: Fattura N°, Data Fat, Cliente Fat, Importo Fat, MRN, Data MRN, Cliente Dog, Importo Dog,
    #         Score Fat, Score Data, Score Import, Score Client, Score Paese, Score Totale, Classe, Spiegazione
    widths = [14, 12, 28, 14, 24, 12, 28, 14, 10, 10, 10, 10, 10, 12, 14, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# FASTAPI APP
# ════════════════════════════════════════════════════════════════════════════
app = FastAPI(title="SIRIO - Sistema di Riconciliazione", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["*"], allow_headers=["*"])


# ── Endpoint: gestione profili per utente ────────────────────────────────────
@app.get("/api/profiles")
async def get_user_profiles_list(token: str = Depends(require_token)):
    """Ritorna i profili dell'utente corrente (autenticato)."""
    user_profiles = get_user_profiles(token)
    return user_profiles


@app.post("/api/profiles/{profile_name}")
async def save_user_profile(
    profile_name: str,
    col_mapping: dict,
    token: str = Depends(require_token)
):
    """Salva un profilo per l'utente corrente."""
    success = save_profile(token, profile_name, col_mapping)
    if success:
        return {"status": "ok", "message": f"Profilo '{profile_name}' salvato"}
    else:
        raise HTTPException(500, f"Errore salvataggio profilo '{profile_name}'")


@app.delete("/api/profiles/{profile_name}")
async def delete_user_profile(
    profile_name: str,
    token: str = Depends(require_token)
):
    """Elimina un profilo dell'utente."""
    success = delete_profile(token, profile_name)
    if success:
        return {"status": "ok", "message": f"Profilo '{profile_name}' eliminato"}
    else:
        raise HTTPException(500, f"Errore eliminazione profilo '{profile_name}'")


# ── Endpoint: stato token ────────────────────────────────────────────────────
@app.get("/api/me")
async def get_me(token: str = Depends(require_token)):
    info = TOKENS[token]
    used = USAGE.get(token, 0)
    max_rows = info.get("max_rows", 0)
    return {
        "name":     info.get("name"),
        "rows_used": used,
        "max_rows": max_rows,
        "remaining": (max_rows - used) if max_rows > 0 else "∞",
        "active":   info.get("active", True),
    }


# ── Endpoint: anteprima colonne ──────────────────────────────────────────────
@app.post("/api/preview")
async def api_preview(
    file_dogana:  UploadFile = File(...),
    file_fatture: UploadFile = File(...),
    token: str = Depends(require_token),
):
    """Legge solo gli header dei file, non consuma quota."""
    dog_data = await file_dogana.read()
    fat_data = await file_fatture.read()

    dog_rows = read_excel_rows(dog_data, file_dogana.filename)
    fat_rows = read_excel_rows(fat_data, file_fatture.filename)

    if not dog_rows or not fat_rows:
        raise HTTPException(400, "Impossibile leggere i file")

    return {
        "dog_headers": list(dog_rows[0].keys()),
        "fat_headers": list(fat_rows[0].keys()),
    }


# ── Endpoint: riconciliazione ────────────────────────────────────────────────
@app.post("/api/reconcile")
async def api_reconcile(
    request: Request,
    file_dogana:  UploadFile = File(...),
    file_fatture: UploadFile = File(...),
    col_mapping:  str = Form(None),
    token: str = Depends(require_token),
):
    dog_data = await file_dogana.read()
    fat_data = await file_fatture.read()

    dog_rows = read_excel_rows(dog_data, file_dogana.filename)
    fat_rows = read_excel_rows(fat_data, file_fatture.filename)

    # Consuma righe dal token (numero di fatture elaborate)
    num_fat_rows = len(fat_rows)  # Includes header row
    consume_rows(token, num_fat_rows)

    if not dog_rows or not fat_rows:
        raise HTTPException(400, "Impossibile leggere i file — controlla il formato (supportati: .xlsx, .xls, .csv)")

    dog_headers = list(dog_rows[0].keys())
    fat_headers = list(fat_rows[0].keys())

    # Parse mapping se fornito, altrimenti auto-detect
    col_map = None
    if col_mapping:
        try:
            col_map = json.loads(col_mapping)
        except:
            pass

    results = reconcile(fat_rows, dog_rows, fat_headers, dog_headers, col_map)

    # Statistiche
    n_certo    = sum(1 for r in results if r["candidati"] and r["candidati"][0]["class"] == "CERTO")
    n_prob     = sum(1 for r in results if r["candidati"] and r["candidati"][0]["class"] == "PROBABILE")
    n_poss     = sum(1 for r in results if r["candidati"] and r["candidati"][0]["class"] == "POSSIBILE")
    n_debole   = sum(1 for r in results if r["candidati"] and r["candidati"][0]["class"] == "DEBOLE")
    n_nessuno  = sum(1 for r in results if not r["candidati"])

    write_log(token, "reconcile", {
        "file_dogana":  file_dogana.filename,
        "file_fatture": file_fatture.filename,
        "n_fatture":    len(results),
        "n_mrn":        len(dog_rows),
        "stats": {
            "CERTO": n_certo, "PROBABILE": n_prob,
            "POSSIBILE": n_poss, "DEBOLE": n_debole, "NO_MATCH": n_nessuno
        }
    })

    log.info(f"[{TOKENS[token]['name']}] Riconciliazione: {len(fat_rows)} fatture × {len(dog_rows)} MRN "
             f"→ CERTO:{n_certo} PROB:{n_prob} POSS:{n_poss} DEB:{n_debole} NO:{n_nessuno}")

    # Serializza (converti date/oggetti non serializzabili)
    def _serial(obj):
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        return str(obj)

    return JSONResponse(json.loads(json.dumps({
        "results": results,
        "stats": {
            "n_fatture":  len(results),
            "n_mrn":      len(dog_rows),
            "CERTO":      n_certo,
            "PROBABILE":  n_prob,
            "POSSIBILE":  n_poss,
            "DEBOLE":     n_debole,
            "NESSUN_MATCH": n_nessuno,
        }
    }, default=_serial)))


# ── Endpoint: export Excel ───────────────────────────────────────────────────
@app.post("/api/export")
async def api_export(
    request: Request,
    file_dogana:  UploadFile = File(...),
    file_fatture: UploadFile = File(...),
    col_mapping:  str = Form(None),
    token: str = Depends(require_token),
):
    dog_data = await file_dogana.read()
    fat_data = await file_fatture.read()
    dog_rows = read_excel_rows(dog_data, file_dogana.filename)
    fat_rows = read_excel_rows(fat_data, file_fatture.filename)

    if not dog_rows or not fat_rows:
        raise HTTPException(400, "File Excel non leggibili")

    # Consuma righe dal token (numero di fatture elaborate)
    num_fat_rows = len(fat_rows)
    consume_rows(token, num_fat_rows)

    dog_headers = list(dog_rows[0].keys())
    fat_headers = list(fat_rows[0].keys())

    # Parse mapping se fornito
    col_map = None
    if col_mapping:
        try:
            col_map = json.loads(col_mapping)
        except:
            pass

    results = reconcile(fat_rows, dog_rows, fat_headers, dog_headers, col_map)
    xlsx_bytes = build_output_excel(results)

    fname = f"Riconciliazione_MRN_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    write_log(token, "export_excel", {
        "filename": fname,
        "n_fatture": len(results),
        "n_mrn": len(dog_rows),
    })

    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ── ADMIN: gestione token ────────────────────────────────────────────────────
@app.get("/api/admin/tokens")
async def admin_tokens(token: str = Depends(require_token)):
    if TOKENS[token].get("name") != "SIRIO_ADMIN":
        raise HTTPException(403, "Riservato agli amministratori")
    result = []
    for t, info in TOKENS.items():
        used = USAGE.get(t, 0)
        result.append({
            "token":     t,
            "name":      info.get("name"),
            "max_rows":  info.get("max_rows", 0),
            "rows_used": used,
            "remaining": (info["max_rows"] - used) if info.get("max_rows", 0) > 0 else "∞",
            "active":    info.get("active", True),
            "created_at": info.get("created_at", ""),
            "notes":     info.get("notes", ""),
        })
    return result

@app.post("/api/admin/tokens")
async def admin_create_token(
    request: Request,
    token: str = Depends(require_token),
):
    if TOKENS[token].get("name") != "SIRIO_ADMIN":
        raise HTTPException(403, "Riservato agli amministratori")
    body = await request.json()
    name     = body.get("name", "Cliente")
    max_rows = int(body.get("max_rows", 1000))
    notes    = body.get("notes", "")
    new_tok  = secrets.token_urlsafe(20)
    TOKENS[new_tok] = {
        "name":       name,
        "max_rows":   max_rows,
        "active":     True,
        "created_at": datetime.now().isoformat(),
        "notes":      notes,
    }
    _save_tokens(TOKENS)
    write_log(token, "create_token", {"new_token_name": name, "max_rows": max_rows})
    log.info(f"Token creato: {name} ({max_rows} righe) — {new_tok}")
    return {"token": new_tok, "name": name, "max_rows": max_rows}

@app.patch("/api/admin/tokens/{target_token}")
async def admin_update_token(
    target_token: str,
    request: Request,
    token: str = Depends(require_token),
):
    if TOKENS[token].get("name") != "SIRIO_ADMIN":
        raise HTTPException(403, "Riservato agli amministratori")
    if target_token not in TOKENS:
        raise HTTPException(404, "Token non trovato")
    body = await request.json()

    # Se si richiede cambio token
    if "token" in body:
        new_token = body["token"].strip()
        if not new_token:
            raise HTTPException(400, "Token non può essere vuoto")
        if new_token in TOKENS:
            raise HTTPException(400, "Token già esiste")
        # Copia i dati al nuovo token
        TOKENS[new_token] = TOKENS[target_token]
        # Copia anche l'usage
        if target_token in USAGE:
            USAGE[new_token] = USAGE[target_token]
        # Elimina il vecchio
        del TOKENS[target_token]
        if target_token in USAGE:
            del USAGE[target_token]
        # Salva e log
        _save_tokens(TOKENS)
        write_log(token, "update_token", {"target": target_token, "new_token": new_token, "action": "token_changed"})
        return {"ok": True, "new_token": new_token}

    # Aggiornamenti normali
    if "active" in body:
        TOKENS[target_token]["active"] = bool(body["active"])
    if "max_rows" in body:
        TOKENS[target_token]["max_rows"] = int(body["max_rows"])
    if "name" in body:
        TOKENS[target_token]["name"] = body["name"]
    _save_tokens(TOKENS)
    write_log(token, "update_token", {"target": target_token, "changes": body})
    return {"ok": True}


# ── ADMIN: log attività ──────────────────────────────────────────────────────
@app.get("/api/admin/logs")
async def admin_logs(
    token: str = Depends(require_token),
    limit: int = 100,
):
    if TOKENS[token].get("name") != "SIRIO_ADMIN":
        raise HTTPException(403, "Riservato agli amministratori")
    if not LOGS_FILE.exists():
        return []
    lines = LOGS_FILE.read_text(encoding="utf-8").strip().splitlines()
    entries = []
    for line in lines[-limit:]:
        try:
            entries.append(json.loads(line))
        except:
            pass
    return list(reversed(entries))


# ════════════════════════════════════════════════════════════════════════════
# FRONTEND — HTML inline (singolo modulo)
# ════════════════════════════════════════════════════════════════════════════
HTML = r"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>SIRIO · Sistema di Riconciliazione</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
  :root {
    --bg:      #0f1117;
    --surface: #1a1d2e;
    --card:    #212438;
    --border:  #2d3158;
    --accent:  #4f7fff;
    --accent2: #00d4aa;
    --text:    #e2e8f0;
    --muted:   #7b84a3;
    --certo:   #00C851;
    --prob:    #33B5E5;
    --poss:    #FFBB33;
    --debole:  #FF4444;
    --none:    #888;
    --radius:  10px;
    --font:    'Segoe UI', system-ui, sans-serif;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { background: var(--bg); color: var(--text); font-family: var(--font); min-height: 100vh; }

  /* ── TOP BAR ── */
  .topbar {
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    padding: 0 24px;
    height: 56px;
    display: flex; align-items: center; justify-content: space-between;
    position: sticky; top: 0; z-index: 100;
  }
  .logo { font-size: 1.1rem; font-weight: 700; letter-spacing: .5px; color: var(--text); }
  .logo span { color: var(--accent); }
  .user-chip {
    display: flex; align-items: center; gap: 10px;
    background: var(--card); border: 1px solid var(--border);
    border-radius: 20px; padding: 4px 14px; font-size: .85rem;
  }
  .quota-bar { width: 80px; height: 6px; background: var(--border); border-radius: 3px; overflow: hidden; }
  .quota-fill { height: 100%; background: var(--accent2); border-radius: 3px; transition: width .4s; }

  /* ── LAYOUT ── */
  .wrap { max-width: 1100px; margin: 0 auto; padding: 24px 20px; }

  /* ── TABS ── */
  .tabs { display: flex; gap: 4px; margin-bottom: 24px; border-bottom: 1px solid var(--border); }
  .tab-btn {
    background: none; border: none; color: var(--muted); padding: 10px 18px;
    font-size: .9rem; cursor: pointer; border-bottom: 2px solid transparent;
    transition: all .2s; margin-bottom: -1px; font-family: var(--font);
  }
  .tab-btn.active { color: var(--accent); border-bottom-color: var(--accent); font-weight: 600; }
  .tab-btn:hover:not(.active) { color: var(--text); }
  .tab-pane { display: none; }
  .tab-pane.active { display: block; }

  /* ── CARDS ── */
  .card {
    background: var(--card); border: 1px solid var(--border);
    border-radius: var(--radius); padding: 20px; margin-bottom: 16px;
  }
  .card h3 { font-size: .95rem; color: var(--muted); text-transform: uppercase;
              letter-spacing: .8px; margin-bottom: 14px; }

  /* ── DROP ZONES ── */
  .dropzone {
    border: 2px dashed var(--border); border-radius: var(--radius);
    padding: 28px 20px; text-align: center; cursor: pointer;
    transition: all .2s; position: relative;
  }
  .dropzone:hover, .dropzone.over { border-color: var(--accent); background: rgba(79,127,255,.06); }
  .dropzone input[type=file] {
    position: absolute; inset: 0; opacity: 0; cursor: pointer; width: 100%; height: 100%;
  }
  .dropzone .dz-icon { font-size: 2rem; margin-bottom: 8px; }
  .dropzone .dz-label { font-size: .9rem; color: var(--muted); }
  .dropzone .dz-label strong { color: var(--accent); }
  .dropzone.loaded { border-color: var(--accent2); background: rgba(0,212,170,.05); }
  .dropzone.loaded .dz-icon::after { content: ' ✓'; color: var(--accent2); }

  /* ── BUTTONS ── */
  .btn {
    display: inline-flex; align-items: center; gap: 7px;
    padding: 10px 20px; border-radius: 7px; border: none;
    font-size: .9rem; font-weight: 600; cursor: pointer;
    font-family: var(--font); transition: all .15s;
  }
  .btn-primary { background: var(--accent); color: #fff; }
  .btn-primary:hover { background: #3d6aff; }
  .btn-success { background: var(--accent2); color: #0f1117; }
  .btn-success:hover { filter: brightness(1.1); }
  .btn-ghost { background: var(--border); color: var(--text); }
  .btn-ghost:hover { background: #3a3f60; }
  .btn:disabled { opacity: .4; cursor: not-allowed; }

  /* ── STATS ── */
  .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(120px, 1fr)); gap: 12px; }
  .stat-box { background: var(--surface); border: 1px solid var(--border); border-radius: 8px;
               padding: 14px; text-align: center; }
  .stat-box .num { font-size: 1.8rem; font-weight: 700; }
  .stat-box .lbl { font-size: .75rem; color: var(--muted); margin-top: 2px; }
  .stat-certo   .num { color: var(--certo); }
  .stat-prob    .num { color: var(--prob); }
  .stat-poss    .num { color: var(--poss); }
  .stat-debole  .num { color: var(--debole); }
  .stat-none    .num { color: var(--none); }
  .stat-tot     .num { color: var(--text); }

  /* ── RISULTATI ── */
  .fat-block {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 8px; margin-bottom: 10px; overflow: hidden;
  }
  .fat-header {
    padding: 12px 16px; display: flex; align-items: center; gap: 14px;
    cursor: pointer; transition: background .15s;
  }
  .fat-header:hover { background: var(--card); }
  .fat-info { flex: 1; }
  .fat-num { font-weight: 600; font-size: .95rem; }
  .fat-meta { font-size: .8rem; color: var(--muted); margin-top: 2px; }
  .fat-badge {
    padding: 3px 10px; border-radius: 12px; font-size: .75rem;
    font-weight: 700; letter-spacing: .5px;
  }
  .badge-CERTO      { background: rgba(0,200,81,.15);  color: var(--certo); }
  .badge-PROBABILE  { background: rgba(51,181,229,.15); color: var(--prob); }
  .badge-POSSIBILE  { background: rgba(255,187,51,.15); color: var(--poss); }
  .badge-DEBOLE     { background: rgba(255,68,68,.15);  color: var(--debole); }
  .badge-NESSUN\ MATCH { background: rgba(136,136,136,.15); color: var(--none); }

  .fat-candidates { border-top: 1px solid var(--border); display: none; }
  .fat-candidates.open { display: block; }
  .candidate-row {
    padding: 10px 16px; border-bottom: 1px solid var(--border);
    display: grid; grid-template-columns: 24px 220px 70px 110px 1fr;
    gap: 10px; align-items: start; font-size: .85rem;
  }
  .candidate-row:last-child { border-bottom: none; }
  .cand-rank { color: var(--muted); font-size: .75rem; padding-top: 2px; }
  .cand-mrn  { font-family: monospace; font-size: .82rem; color: var(--accent2); }
  .cand-score { font-weight: 700; text-align: center; }
  .cand-class { text-align: center; }
  .cand-notes { color: var(--muted); font-size: .78rem; line-height: 1.4; }
  .cand-method { font-size: .7rem; color: var(--muted); margin-top: 3px; font-style: italic; }

  /* ── TOKEN INPUT ── */
  .token-gate {
    position: fixed; inset: 0; background: var(--bg); z-index: 999;
    display: flex; align-items: center; justify-content: center;
  }
  .token-box {
    background: var(--card); border: 1px solid var(--border);
    border-radius: 14px; padding: 40px; width: 380px; text-align: center;
  }
  .token-box h2 { margin-bottom: 8px; }
  .token-box p  { color: var(--muted); font-size: .9rem; margin-bottom: 24px; }
  .token-box input {
    width: 100%; padding: 10px 14px; background: var(--surface);
    border: 1px solid var(--border); border-radius: 7px;
    color: var(--text); font-size: .95rem; margin-bottom: 12px;
    font-family: monospace;
  }
  .token-box input:focus { outline: none; border-color: var(--accent); }
  .token-err { color: var(--debole); font-size: .85rem; margin-top: 8px; }

  /* ── ADMIN ── */
  .admin-table { width: 100%; border-collapse: collapse; font-size: .75rem; }
  .admin-table th { background: var(--surface); color: var(--muted); text-align: left;
                     padding: 6px 8px; font-weight: 600; border-bottom: 1px solid var(--border); font-size: .75rem; }
  .admin-table td { padding: 6px 8px; border-bottom: 1px solid var(--border); }
  .admin-table tr:hover td { background: var(--surface); }
  .tok-cell { font-family: monospace; font-size: .7rem; color: var(--muted); cursor: pointer; word-break: break-all; }
  .tok-cell:hover { color: var(--accent2); }

  .log-entry { display: grid; grid-template-columns: 170px 70px 90px 1fr;
               gap: 8px; padding: 5px 0; border-bottom: 1px solid var(--border);
               font-size: .75rem; align-items: start; }
  .log-ts { color: var(--muted); }
  .log-user { color: var(--accent); font-weight: 600; }
  .log-action { color: var(--accent2); }
  .log-detail { color: var(--muted); }

  .form-row { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-bottom: 12px; }
  .form-group label { font-size: .8rem; color: var(--muted); display: block; margin-bottom: 4px; }
  .form-group input {
    width: 100%; padding: 8px 12px; background: var(--surface);
    border: 1px solid var(--border); border-radius: 6px;
    color: var(--text); font-family: var(--font); font-size: .9rem;
  }
  .form-group input:focus { outline: none; border-color: var(--accent); }

  .spinner { display: inline-block; width: 16px; height: 16px; border: 2px solid rgba(255,255,255,.3);
             border-top-color: #fff; border-radius: 50%; animation: spin .7s linear infinite; }
  @keyframes spin { to { transform: rotate(360deg); } }

  .alert { padding: 10px 14px; border-radius: 7px; font-size: .85rem; margin-bottom: 12px; }
  .alert-error   { background: rgba(255,68,68,.12);   color: var(--debole); border: 1px solid rgba(255,68,68,.3); }
  .alert-success { background: rgba(0,200,81,.1);    color: var(--certo);  border: 1px solid rgba(0,200,81,.3); }

  /* ═══════════════════════════════════════════════════════════════════════════ */
  /* ── IVISTO CC599C VIEWER (scopato su #ivisto-panel) ── */
  /* ═════════════════════════════════════════════════════════════════════════════ */

  #ivisto-panel {
    --bg: #0f1117; --surface: #181c27; --surface2: #1e2336; --border: #2a3050; --border2: #3a4570;
    --accent: #4f8ef7; --accent2: #2ecc8a; --warn: #f5a623; --danger: #e05555;
    --text: #e8eaf0; --text2: #8890a8; --text3: #5a6280;
    --mono: 'IBM Plex Mono', monospace; --sans: 'IBM Plex Sans', sans-serif;
  }

  /* MAIN */
  #ivisto-panel .main { max-width: 860px; margin: 0 auto; padding: 24px; }

  /* CARD */
  #ivisto-panel .card { background: var(--surface); border: 1px solid var(--border); border-radius: 8px; padding: 20px; margin-bottom: 16px; }
  #ivisto-panel .card-head { font-family: var(--mono); font-size: 11px; color: var(--text3); letter-spacing: 1px; text-transform: uppercase; margin-bottom: 16px; padding-bottom: 10px; border-bottom: 1px solid var(--border); display: flex; justify-content: space-between; align-items: center; }

  /* UPLOAD */
  #ivisto-panel .upload-zone { border: 1px dashed var(--border2); border-radius: 8px; padding: 40px 20px; text-align: center; cursor: pointer; transition: border-color 0.15s, background 0.15s; background: var(--surface2); }
  #ivisto-panel .upload-zone:hover { border-color: var(--accent); background: rgba(79,142,247,0.05); }
  #ivisto-panel .upload-zone input { display: none; }
  #ivisto-panel .upload-title { font-size: 13px; font-weight: 500; color: var(--text); margin-bottom: 6px; }
  #ivisto-panel .upload-sub { font-size: 11px; color: var(--text3); font-family: var(--mono); }

  /* TAB SWITCHER */
  #ivisto-panel .tab-bar { display: flex; gap: 0; margin-bottom: 14px; border-bottom: 1px solid var(--border); }
  #ivisto-panel .tab-btn { padding: 8px 16px; font-size: 12px; color: var(--text2); cursor: pointer; background: none; border: none; border-bottom: 2px solid transparent; font-family: var(--sans); transition: all 0.15s; margin-bottom: -1px; }
  #ivisto-panel .tab-btn:hover { color: var(--text); }
  #ivisto-panel .tab-btn.active { color: var(--accent); border-bottom-color: var(--accent); }
  #ivisto-panel .tab-panel { display: none; }
  #ivisto-panel .tab-panel.active { display: block; }

  /* TEXTAREA XML */
  #ivisto-panel .xml-input { width: 100%; background: var(--surface2); border: 1px solid var(--border); border-radius: 6px; color: var(--text); padding: 10px 12px; font-size: 11px; font-family: var(--mono); resize: vertical; min-height: 140px; line-height: 1.6; transition: border-color 0.15s; }
  #ivisto-panel .xml-input:focus { outline: none; border-color: var(--accent); }
  #ivisto-panel .xml-input::placeholder { color: var(--text3); }
  #ivisto-panel .paste-toolbar { display: flex; justify-content: flex-end; gap: 8px; margin-top: 10px; }

  /* BOTTONI */
  #ivisto-panel .btn { padding: 7px 14px; border-radius: 5px; font-size: 12px; font-weight: 500; cursor: pointer; border: none; font-family: var(--sans); transition: all 0.15s; display: inline-flex; align-items: center; gap: 5px; }
  #ivisto-panel .btn-ghost { background: transparent; color: var(--text2); border: 1px solid var(--border2); }
  #ivisto-panel .btn-ghost:hover { border-color: var(--accent); color: var(--accent); }
  #ivisto-panel .btn-primary { background: var(--accent); color: #fff; }
  #ivisto-panel .btn-primary:hover { background: #6a9ff8; }
  #ivisto-panel .toolbar { display: flex; gap: 8px; justify-content: flex-end; margin-top: 16px; }

  /* ALERT */
  #ivisto-panel .alert { padding: 10px 14px; border-radius: 5px; font-size: 12px; margin-bottom: 16px; display: flex; align-items: flex-start; gap: 8px; line-height: 1.6; }
  #ivisto-panel .alert-err { background: rgba(224,85,85,0.1); border: 1px solid rgba(224,85,85,0.25); color: var(--danger); }
  #ivisto-panel .alert-err ul { margin: 4px 0 0 16px; }

  /* RESULT BLOCK */
  #ivisto-panel .result-block { border-radius: 6px; padding: 14px 16px; border-left: 3px solid; margin-bottom: 16px; }
  #ivisto-panel .result-A1 { background: rgba(79,142,247,0.08); border-color: var(--accent); }
  #ivisto-panel .result-A1 .result-code, #ivisto-panel .result-A1 .result-desc { color: #7fb3ff; }
  #ivisto-panel .result-A2 { background: rgba(46,204,138,0.08); border-color: var(--accent2); }
  #ivisto-panel .result-A2 .result-code, #ivisto-panel .result-A2 .result-desc { color: var(--accent2); }
  #ivisto-panel .result-A3 { background: rgba(245,166,35,0.08); border-color: var(--warn); }
  #ivisto-panel .result-A3 .result-code, #ivisto-panel .result-A3 .result-desc { color: var(--warn); }
  #ivisto-panel .result-B  { background: rgba(224,85,85,0.08); border-color: var(--danger); }
  #ivisto-panel .result-B  .result-code, #ivisto-panel .result-B  .result-desc { color: var(--danger); }
  #ivisto-panel .result-default { background: var(--surface2); border-color: var(--border2); }
  #ivisto-panel .result-code { font-family: var(--mono); font-size: 15px; font-weight: 600; margin-bottom: 5px; }
  #ivisto-panel .result-desc { font-size: 12px; line-height: 1.6; opacity: 0.85; }

  /* BADGE */
  #ivisto-panel .badge { display: inline-flex; align-items: center; gap: 5px; border-radius: 10px; font-size: 10px; font-family: var(--mono); padding: 2px 8px; font-weight: 600; }
  #ivisto-panel .badge-ok      { background: rgba(46,204,138,0.15); color: var(--accent2); }
  #ivisto-panel .badge-info    { background: rgba(79,142,247,0.15); color: var(--accent); }
  #ivisto-panel .badge-warn    { background: rgba(245,166,35,0.15); color: var(--warn); }
  #ivisto-panel .badge-err     { background: rgba(224,85,85,0.15);  color: var(--danger); }
  #ivisto-panel .badge-neutral { background: var(--surface2); color: var(--text3); border: 1px solid var(--border); }

  /* DOT */
  #ivisto-panel .dot { display: inline-block; width: 6px; height: 6px; border-radius: 50%; flex-shrink: 0; }
  #ivisto-panel .dot-ok      { background: var(--accent2); }
  #ivisto-panel .dot-info    { background: var(--accent); }
  #ivisto-panel .dot-warn    { background: var(--warn); }
  #ivisto-panel .dot-danger  { background: var(--danger); }
  #ivisto-panel .dot-neutral { background: var(--text3); }

  /* STAT BOXES */
  #ivisto-panel .stat-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
  #ivisto-panel .stat-box { background: var(--surface2); border: 1px solid var(--border); border-radius: 6px; padding: 12px 14px; }
  #ivisto-panel .stat-box.full { grid-column: 1 / -1; }
  #ivisto-panel .stat-label { font-size: 10px; color: var(--text3); font-family: var(--mono); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 5px; }
  #ivisto-panel .stat-value { font-size: 13px; font-weight: 500; color: var(--text); line-height: 1.4; word-break: break-all; }
  #ivisto-panel .stat-value.mono { font-family: var(--mono); font-size: 13px; color: var(--accent); }
  #ivisto-panel .stat-sub { font-family: var(--mono); font-size: 10px; color: var(--text3); margin-top: 3px; }

  /* TAG non in rubrica */
  #ivisto-panel .tag-unknown { display: inline-block; background: rgba(224,85,85,0.1); border: 1px solid rgba(224,85,85,0.2); border-radius: 3px; font-size: 9px; padding: 0 4px; color: var(--danger); margin-left: 5px; vertical-align: middle; font-family: var(--mono); }

  /* SEZIONI DOC */
  #ivisto-panel .doc-section { margin-bottom: 20px; }
  #ivisto-panel .doc-section-title { font-family: var(--mono); font-size: 10px; color: var(--text3); text-transform: uppercase; letter-spacing: 1px; margin-bottom: 10px; padding-bottom: 8px; border-bottom: 1px solid var(--border); }

  /* NOTE */
  #ivisto-panel .note-text { font-size: 11px; color: var(--text3); line-height: 1.7; }

  /* PRINT */
  #ivisto-panel #printDate { display: none; }

  @media print {
    #ivisto-panel .main { max-width: 100%; padding: 0; }
    #ivisto-panel .card { background: #fff; border: 1px solid #ccc; padding: 12px; margin-bottom: 10px; box-shadow: none; }
    #ivisto-panel .card-head { color: #555; border-bottom-color: #ccc; margin-bottom: 10px; padding-bottom: 6px; }
    #ivisto-panel .stat-box { background: #f5f5f5; border-color: #ddd; padding: 8px 10px; }
    #ivisto-panel .stat-label { color: #666; }
    #ivisto-panel .stat-value { color: #111; font-size: 11px; }
    #ivisto-panel .stat-value.mono { color: #1a4d8f; }
    #ivisto-panel .stat-sub { color: #888; }
    #ivisto-panel .stat-grid { gap: 6px; }
    #ivisto-panel .result-block { padding: 10px 12px; margin-bottom: 10px; }
    #ivisto-panel .result-code { font-size: 13px; }
    #ivisto-panel .result-desc { font-size: 10px; opacity: 1; }
    #ivisto-panel .result-A2 { background: #e8f7ee !important; border-color: #2a9e6a !important; }
    #ivisto-panel .result-A2 .result-code, #ivisto-panel .result-A2 .result-desc { color: #1a5c3a !important; }
    #ivisto-panel .result-A1 { background: #e8eef9 !important; border-color: #3a6ec0 !important; }
    #ivisto-panel .result-A1 .result-code, #ivisto-panel .result-A1 .result-desc { color: #1a3d8f !important; }
    #ivisto-panel .result-A3 { background: #fdf3e3 !important; border-color: #c07a10 !important; }
    #ivisto-panel .result-A3 .result-code, #ivisto-panel .result-A3 .result-desc { color: #7a4d00 !important; }
    #ivisto-panel .result-B  { background: #fdeaea !important; border-color: #b03030 !important; }
    #ivisto-panel .result-B  .result-code, #ivisto-panel .result-B  .result-desc { color: #8f1a1a !important; }
    #ivisto-panel .doc-section { margin-bottom: 12px; }
    #ivisto-panel .doc-section-title { font-size: 9px; color: #888; border-bottom-color: #ccc; padding-bottom: 4px; margin-bottom: 6px; }
    #ivisto-panel .badge { font-size: 9px; }
    #ivisto-panel .tag-unknown { background: #fdeaea; color: #c00; border-color: #faa; }
    #ivisto-panel .note-text { font-size: 9px; color: #888; }
    #ivisto-panel #printDate { display: block !important; font-size: 10px; color: #888; margin-top: 6px; font-family: monospace; }
    #ivisto-panel .toolbar, #ivisto-panel #uploadCard { display: none !important; }
  }
</style>
</head>
<body>

<!-- TOKEN GATE -->
<div class="token-gate" id="gate">
  <div class="token-box">
    <div style="font-size:2.5rem;margin-bottom:12px">🔐</div>
    <h2>SIRIO · Sistema di Riconciliazione</h2>
    <p>Inserisci il token di accesso</p>
    <input type="text" id="tokenInput" placeholder="Incolla il token..." autocomplete="off" spellcheck="false">

    <!-- INFORMATIVA PRIVACY -->
    <div style="margin-top:20px;padding:12px;background:rgba(255,215,0,0.05);border-left:3px solid #ffd700;border-radius:4px;font-size:.85rem">
      <div style="cursor:pointer;display:flex;justify-content:space-between;align-items:center" onclick="togglePrivacyInfo()">
        <strong>📋 Informativa sulla Privacy</strong>
        <span id="privacyToggle">▼</span>
      </div>
      <div id="privacyInfo" style="display:none;margin-top:10px;max-height:200px;overflow-y:auto;line-height:1.4;color:var(--text-secondary)">
        <p style="margin:0 0 8px 0"><strong>Trattamento dei Dati Sensibili</strong></p>
        <p style="margin:0 0 8px 0">
          Accedendo a SIRIO, confermi di autorizzare il trattamento dei tuoi dati sensibili relativi a fatture e dichiarazioni doganali, in conformità al Regolamento (UE) 2016/679 (GDPR).
        </p>
        <p style="margin:0 0 8px 0"><strong>Protezione dei Dati</strong></p>
        <p style="margin:0 0 8px 0">
          <strong>I dati caricati NON verranno conservati sul nostro sistema.</strong> I file sono elaborati esclusivamente per la riconciliazione e vengono eliminati al termine dell'elaborazione. Non conserviamo copie, backup o tracce dei tuoi dati.
        </p>
        <p style="margin:0 0 8px 0"><strong>Responsabilità</strong></p>
        <p style="margin:0 0 8px 0">
          L'utente rimane il titolare dei propri dati e mantiene piena responsabilità sulla liceità del loro trattamento. Accettando, dichiari di avere l'autorità per fornire questi dati.
        </p>
      </div>
    </div>

    <!-- CHECKBOX ACCETTAZIONE -->
    <div style="margin-top:12px;display:flex;align-items:center;gap:8px">
      <input type="checkbox" id="privacyCheckbox" onchange="updateAccessButton()">
      <label for="privacyCheckbox" style="margin:0;font-size:.9rem;cursor:pointer">
        Accetto l'informativa sulla privacy e il trattamento dei miei dati sensibili
      </label>
    </div>

    <button class="btn btn-primary" style="width:100%;margin-top:12px" id="accessBtn" onclick="submitToken()" disabled>Accedi</button>
    <div class="token-err" id="tokenErr"></div>
  </div>
</div>

<!-- TOP BAR -->
<div class="topbar" id="topbar" style="display:none">
  <div class="logo">SIRIO · <span>Sistema di Riconciliazione</span></div>
  <div style="display:flex;align-items:center;gap:16px">
    <div class="user-chip">
      <span id="userName">—</span>
      <div class="quota-bar"><div class="quota-fill" id="quotaFill" style="width:0%"></div></div>
      <span id="quotaText" style="font-size:.78rem;color:var(--muted)"></span>
    </div>
    <button class="btn btn-ghost" onclick="logout()" style="font-size:.85rem;padding:6px 12px">🚪 Disconnetti</button>
  </div>
</div>

<!-- MAIN -->
<div class="wrap" id="main" style="display:none">

  <!-- TABS -->
  <div class="tabs">
    <button class="tab-btn active" onclick="showTab('tab-rec', this)">Riconciliazione</button>
    <button class="tab-btn" onclick="showTab('tab-ivisto', this)">IVISTO CC599C</button>
    <button class="tab-btn" onclick="showTab('tab-admin', this)" id="adminTabBtn" style="display:none">Amministrazione</button>
  </div>

  <!-- TAB: RICONCILIAZIONE -->
  <div class="tab-pane active" id="tab-rec">

    <!-- Upload -->
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">
      <div class="card">
        <h3>📦 File Estratto dal Cassetto Doganale</h3>
        <div class="dropzone" id="dz-dog" ondragover="dzOver(event,'dz-dog')"
             ondragleave="dzOut('dz-dog')" ondrop="dzDrop(event,'dz-dog','fileDogana')">
          <input type="file" accept=".xlsx,.xls,.csv" onchange="fileSelected(this,'dz-dog','fileDogana','lblDog')">
          <div class="dz-icon">📑</div>
          <div class="dz-label">Trascina qui o <strong>clicca</strong><br>File Excel con MRN e riferimenti fattura</div>
        </div>
        <div id="lblDog" style="margin-top:8px;font-size:.82rem;color:var(--accent2)"></div>
      </div>
      <div class="card">
        <h3>🧾 File Fatture da Contabilità</h3>
        <div class="dropzone" id="dz-fat" ondragover="dzOver(event,'dz-fat')"
             ondragleave="dzOut('dz-fat')" ondrop="dzDrop(event,'dz-fat','fileFatture')">
          <input type="file" accept=".xlsx,.xls,.csv" onchange="fileSelected(this,'dz-fat','fileFatture','lblFat')">
          <div class="dz-icon">🧾</div>
          <div class="dz-label">Trascina qui o <strong>clicca</strong><br>File Excel con elenco fatture cliente</div>
        </div>
        <div id="lblFat" style="margin-top:8px;font-size:.82rem;color:var(--accent2)"></div>
      </div>
    </div>

    <!-- Pannello configurazione colonne -->
    <div id="colConfigPanel" style="display:none;border:1px solid var(--border);border-radius:8px;padding:16px;margin-bottom:16px;background:var(--card-bg)">
      <div style="margin-bottom:16px">
        <label style="font-weight:500;display:block;margin-bottom:6px">📋 Configurazione colonne</label>
        <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
          <select id="selProfile" style="padding:6px 8px;border:1px solid var(--border);border-radius:4px;font-size:.9rem" onchange="loadProfile()">
            <option value="">— Carica profilo salvato —</option>
          </select>
          <input type="text" id="inputProfileName" placeholder="Nome profilo" style="padding:6px 8px;border:1px solid var(--border);border-radius:4px;font-size:.9rem;flex:1;min-width:150px">
          <button class="btn btn-ghost" style="font-size:.85rem;padding:6px 12px" onclick="saveProfile()">💾 Salva</button>
        </div>
      </div>

      <div style="display:grid;grid-template-columns:1fr;gap:16px">
        <!-- Fatture -->
        <div>
          <h4 style="margin:0 0 12px 0;font-size:.95rem;color:var(--accent1)">📄 Fatture Cliente</h4>
          <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:10px">
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Numero Fattura</label>
              <select class="col-select" data-field="fat_num_fattura" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Data Fattura</label>
              <select class="col-select" data-field="fat_data_fattura" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Cliente</label>
              <select class="col-select" data-field="fat_cliente_fat" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Importo</label>
              <select class="col-select" data-field="fat_importo_fat" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Paese</label>
              <select class="col-select" data-field="fat_paese_fat" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
          </div>
        </div>

        <!-- Dogana (collassato) -->
        <details style="border-top:1px solid var(--border);padding-top:12px">
          <summary style="cursor:pointer;font-weight:500;font-size:.9rem;color:var(--accent2);user-select:none">⚙ Configura Dogana (opzionale)</summary>
          <div style="margin-top:12px;display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:10px">
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Codice MRN</label>
              <select class="col-select" data-field="dog_mrn" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Rif. Fattura</label>
              <select class="col-select" data-field="dog_fattura_dog" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Data</label>
              <select class="col-select" data-field="dog_data_dog" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Importo</label>
              <select class="col-select" data-field="dog_importo_dog" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Cliente</label>
              <select class="col-select" data-field="dog_cliente_dog" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
            <div>
              <label style="font-size:.8rem;color:var(--muted);display:block;margin-bottom:4px">Paese</label>
              <select class="col-select" data-field="dog_paese_dog" style="width:100%;padding:6px;border:1px solid var(--border);border-radius:4px;font-size:.85rem"></select>
            </div>
          </div>
        </details>
      </div>
    </div>

    <!-- Azioni -->
    <div style="display:flex;gap:10px;margin-bottom:20px;align-items:center">
      <button class="btn btn-primary" id="btnRec" onclick="doReconcile()" disabled>
        <span class="spinner" id="spinRec" style="display:none"></span>
        ⚖ Avvia Riconciliazione
      </button>
      <button class="btn btn-success" id="btnExport" onclick="doExport()" disabled>
        <span class="spinner" id="spinExp" style="display:none"></span>
        ⬇ Esporta Excel
      </button>
      <div id="alertRec"></div>
    </div>

    <!-- Stats -->
    <div id="statsArea" style="display:none" class="card">
      <h3>Riepilogo</h3>
      <div class="stats-grid">
        <div class="stat-box stat-tot"><div class="num" id="s-tot">0</div><div class="lbl">Fatture</div></div>
        <div class="stat-box stat-tot"><div class="num" id="s-mrn">0</div><div class="lbl">MRN</div></div>
        <div class="stat-box stat-certo"><div class="num" id="s-cert">0</div><div class="lbl">CERTO</div></div>
        <div class="stat-box stat-prob"><div class="num" id="s-prob">0</div><div class="lbl">PROBABILE</div></div>
        <div class="stat-box stat-poss"><div class="num" id="s-poss">0</div><div class="lbl">POSSIBILE</div></div>
        <div class="stat-box stat-debole"><div class="num" id="s-deb">0</div><div class="lbl">DEBOLE</div></div>
        <div class="stat-box stat-none"><div class="num" id="s-none">0</div><div class="lbl">NO MATCH</div></div>
      </div>
    </div>

    <!-- Risultati -->
    <div id="resultsArea"></div>
  </div>

  <!-- TAB: ADMIN -->
  <div class="tab-pane" id="tab-admin">

    <!-- Crea token -->
    <div class="card">
      <h3>➕ Crea Nuovo Token Cliente</h3>
      <div class="form-row">
        <div class="form-group">
          <label>Nome cliente / identificativo</label>
          <input type="text" id="newName" placeholder="es. ACME S.r.l.">
        </div>
        <div class="form-group">
          <label>Numero righe massime</label>
          <input type="number" id="newMaxRows" value="1000" min="1">
        </div>
      </div>
      <div class="form-group" style="margin-bottom:12px">
        <label>Note</label>
        <input type="text" id="newNotes" placeholder="es. Contratto 2025 — max 5000 righe">
      </div>
      <button class="btn btn-primary" onclick="createToken()" title="Crea un nuovo token di accesso per il cliente">Genera Token</button>
      <div id="newTokenResult" style="margin-top:12px;font-family:monospace;color:var(--accent2);font-size:.9rem"></div>
    </div>

    <!-- Lista token -->
    <div class="card">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px">
        <h3 style="margin:0">🔑 Token Attivi</h3>
        <button class="btn btn-ghost" style="font-size:.8rem;padding:6px 12px" onclick="loadTokens()" title="Ricarica l'elenco dei token attivi">↻ Aggiorna</button>
      </div>
      <table class="admin-table" id="tokenTable">
        <thead>
          <tr>
            <th>Token</th><th>Nome</th><th>Utilizzi</th><th>Rimanenti</th>
            <th>Stato</th><th>Creato</th><th>Note</th><th>Azioni</th>
          </tr>
        </thead>
        <tbody id="tokenTbody"></tbody>
      </table>
    </div>

    <!-- Log attività -->
    <div class="card">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px">
        <h3 style="margin:0">📋 Log Attività (ultimi 100)</h3>
        <button class="btn btn-ghost" style="font-size:.8rem;padding:6px 12px" onclick="loadLogs()" title="Ricarica il log dell'attività">↻ Aggiorna</button>
      </div>
      <div id="logArea"></div>
    </div>
  </div>

  <!-- TAB: IVISTO CC599C -->
  <div class="tab-pane" id="tab-ivisto">
    <div id="ivisto-panel">
      <div class="main">
        <div id="uploadCard" class="card">
          <div class="card-head">Carica messaggio XML</div>
          <div class="tab-bar">
            <button class="tab-btn active" onclick="ivistoSwitchTab('file')">Carica file</button>
            <button class="tab-btn"        onclick="ivistoSwitchTab('paste')">Incolla XML</button>
          </div>
          <div id="tab-file" class="tab-panel active">
            <div class="upload-zone" id="dropZone" onclick="document.getElementById('fi').click()">
              <div class="upload-title">Seleziona o trascina un file CC599C</div>
              <div class="upload-sub">formato .xml &mdash; struttura CC599C (IVISTO)</div>
              <input type="file" id="fi" accept=".xml,text/xml">
            </div>
          </div>
          <div id="tab-paste" class="tab-panel">
            <textarea id="xmlPaste" class="xml-input" placeholder="Incolla qui il contenuto XML del messaggio CC599C..."></textarea>
            <div class="paste-toolbar">
              <button class="btn btn-ghost" onclick="document.getElementById('xmlPaste').value=''">Cancella</button>
              <button class="btn btn-primary" onclick="ivistoAnalyzeText()">Analizza</button>
            </div>
          </div>
        </div>
        <div id="errBox" style="display:none"></div>
        <div id="out"    style="display:none"></div>
        <div class="toolbar" id="tb" style="display:none">
          <button class="btn btn-ghost"   onclick="ivistoClearDoc()">Nuovo file</button>
          <button class="btn btn-primary" onclick="window.print()">Stampa</button>
        </div>
      </div>
    </div>
  </div>

</div>

<script>
// ── STATE ──────────────────────────────────────────────────────────────────
let TOKEN = localStorage.getItem('mrn_token') || '';
let fileDogana  = null;
let fileFatture = null;
let lastResults = null;
let currentColMapping = null;  // Salva il mapping corrente per l'export
let cachedProfiles = {};  // Cache dei profili caricati dal server

// ── PRIVACY & ACCESS ────────────────────────────────────────────────────────
function togglePrivacyInfo() {
  const info = document.getElementById('privacyInfo');
  const toggle = document.getElementById('privacyToggle');
  if (info.style.display === 'none') {
    info.style.display = 'block';
    toggle.textContent = '▲';
  } else {
    info.style.display = 'none';
    toggle.textContent = '▼';
  }
}

function updateAccessButton() {
  const checkbox = document.getElementById('privacyCheckbox');
  const btn = document.getElementById('accessBtn');
  btn.disabled = !checkbox.checked;
}

// ── AUTH ───────────────────────────────────────────────────────────────────
async function submitToken() {
  const t = document.getElementById('tokenInput').value.trim();
  if (!t) return;
  try {
    const r = await fetch('/api/me', { headers: { 'X-Token': t } });
    if (!r.ok) throw new Error((await r.json()).detail || 'Errore');
    const me = await r.json();
    TOKEN = t;
    localStorage.setItem('mrn_token', t);
    initApp(me);
  } catch(e) {
    document.getElementById('tokenErr').textContent = '✗ ' + e.message;
  }
}
document.getElementById('tokenInput').addEventListener('keydown', e => {
  if (e.key === 'Enter') submitToken();
});

async function initApp(me) {
  document.getElementById('gate').style.display = 'none';
  document.getElementById('topbar').style.display = 'flex';
  document.getElementById('main').style.display = 'block';
  document.getElementById('userName').textContent = me.name;
  updateQuota(me);
  if (me.name === 'SIRIO_ADMIN') {
    document.getElementById('adminTabBtn').style.display = '';
    loadTokens();
    loadLogs();
  }
}

function updateQuota(me) {
  const maxRows = me.max_rows;
  const used = me.rows_used;
  const txt = document.getElementById('quotaText');
  const fill = document.getElementById('quotaFill');
  if (maxRows === 0) {
    txt.textContent = '∞ righe';
    fill.style.width = '10%';
  } else {
    const pct = Math.min(100, (used / maxRows) * 100);
    fill.style.width = pct + '%';
    fill.style.background = pct >= 90 ? 'var(--debole)' : pct >= 70 ? 'var(--poss)' : 'var(--accent2)';
    txt.textContent = `${used}/${maxRows}`;
  }
}

// Auto-login se token salvato
if (TOKEN) {
  fetch('/api/me', { headers: { 'X-Token': TOKEN } })
    .then(r => r.ok ? r.json() : Promise.reject(r))
    .then(me => initApp(me))
    .catch(() => { localStorage.removeItem('mrn_token'); TOKEN = ''; });
}

// ── LOGOUT ─────────────────────────────────────────────────────────────────
function logout() {
  localStorage.removeItem('mrn_token');
  TOKEN = '';
  document.getElementById('gate').style.display = 'flex';
  document.getElementById('topbar').style.display = 'none';
  document.getElementById('main').style.display = 'none';
  document.getElementById('tokenInput').value = '';
  document.getElementById('tokenErr').textContent = '';
  document.getElementById('privacyCheckbox').checked = false;
  updateAccessButton();
}

// ── TABS ───────────────────────────────────────────────────────────────────
function showTab(id, btn) {
  document.querySelectorAll('.tabs ~ .tab-pane').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tabs .tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  btn.classList.add('active');
}

// ── DROP ZONE ──────────────────────────────────────────────────────────────
function dzOver(e, id) { e.preventDefault(); document.getElementById(id).classList.add('over'); }
function dzOut(id) { document.getElementById(id).classList.remove('over'); }
function dzDrop(e, id, storeKey, lblId) {
  e.preventDefault(); dzOut(id);
  const file = e.dataTransfer.files[0];
  if (file) assignFile(file, id, storeKey, lblId);
}
function fileSelected(input, id, storeKey, lblId) {
  if (input.files[0]) assignFile(input.files[0], id, storeKey, lblId);
}
function assignFile(file, dzId, storeKey, lblId) {
  if (storeKey === 'fileDogana') fileDogana = file;
  else if (storeKey === 'fileFatture') fileFatture = file;
  document.getElementById(dzId).classList.add('loaded');
  if (lblId) document.getElementById(lblId).textContent = '✓ ' + file.name;
  checkReady();
}
async function checkReady() {
  const ok = fileDogana && fileFatture;
  if (ok) {
    // Chiamata preview per ottenere le colonne
    const fd = new FormData();
    fd.append('file_dogana', fileDogana);
    fd.append('file_fatture', fileFatture);
    try {
      const r = await fetch('/api/preview', {
        method: 'POST', body: fd, headers: { 'X-Token': TOKEN }
      });
      if (r.ok) {
        const data = await r.json();
        await showColConfigPanel(data.fat_headers, data.dog_headers);
      }
    } catch(e) {
      console.error('Preview error:', e);
    }
  }
  document.getElementById('btnRec').disabled = !ok;
}

async function showColConfigPanel(fatHeaders, dogHeaders) {
  console.log('🎯 showColConfigPanel() avviato');
  const panel = document.getElementById('colConfigPanel');

  if (!panel) {
    console.error('❌ ERRORE: colConfigPanel non trovato nel DOM');
    return;
  }

  panel.style.display = '';

  // Salva i valori attuali prima di ripopolare (se il pannello è già stato configurato)
  const oldValues = {};
  const selectsNodeList = panel.querySelectorAll('.col-select');
  console.log(`🔍 Trovate ${selectsNodeList.length} select.col-select`);

  // Converti NodeList in Array per sicurezza
  const selects = Array.from(selectsNodeList);
  console.log(`✓ Convertite a Array: ${selects.length} elementi`);

  selects.forEach(sel => {
    oldValues[sel.dataset.field] = sel.value;
  });

  // Popola tutte le select con le nuove opzioni di colonna
  selects.forEach(sel => {
    sel.innerHTML = '<option value="">— non usato —</option>';
    fatHeaders.concat(dogHeaders).forEach(h => {
      const opt = document.createElement('option');
      opt.value = h;
      opt.textContent = h;
      sel.appendChild(opt);
    });
    // Ripristina il valore precedente se esiste
    if (oldValues[sel.dataset.field]) {
      sel.value = oldValues[sel.dataset.field];
    }
  });

  // Pre-seleziona i match automatici SOLO se nessun campo è stato già configurato
  const hasConfiguration = selects.length > 0 && selects.some(sel => sel.value !== '');
  if (!hasConfiguration) {
    console.log('🔍 Nessuna configurazione precedente, eseguo autoSelectColumns()');
    autoSelectColumns(fatHeaders, dogHeaders);
  } else {
    console.log('✓ Configurazione precedente trovata, mantengo i valori');
  }

  // NON sovrascrivere currentColMapping se è già stato settato dalla riconciliazione
  if (!currentColMapping || Object.keys(currentColMapping.fat).length === 0) {
    currentColMapping = getColMapping();
  }

  // Popola dropdown profili (asincrono)
  console.log('📝 Carico lista profili...');
  await loadProfileList();
  console.log('✅ Lista profili caricata');

  // CARICA AUTOMATICAMENTE L'ULTIMO PROFILO USATO
  const lastProfile = localStorage.getItem('mrn_last_profile');
  console.log('💾 Ultimo profilo in localStorage:', lastProfile);
  if (lastProfile) {
    const sel = document.getElementById('selProfile');
    console.log('📍 Imposto select value a:', lastProfile);
    sel.value = lastProfile;
    console.log('📲 Chiamo loadProfile()...');
    await loadProfile(lastProfile);
  } else {
    console.log('❌ Nessun ultimo profilo trovato in localStorage');
  }
  console.log('✅ showColConfigPanel() completato');
}

function autoSelectColumns(fatHeaders, dogHeaders) {
  const mapping = { fat: {}, dog: {} };

  // PRIMA COSA: Imposta direttamente "Val. Stat." per importo_dog se esiste
  // NON usare autodetect per questo campo
  const valStatCol = dogHeaders.find(h => h.toLowerCase().includes('val.stat') || h.toLowerCase().includes('val. stat'));
  if (valStatCol) {
    mapping.dog.importo_dog = valStatCol;  // Chiave corretta con underscore
  }

  // Map semplice per auto-rilevamento
  const simpleMap = {
    'num_fattura': ['numero', 'numero fattura', 'n.', 'invoice', 'numero_fattura'],
    'data_fattura': ['data', 'data fattura', 'date', 'data_fattura'],
    'cliente_fat': ['cliente', 'cliente fattura', 'ragione', 'customer', 'cliente_fat'],
    'importo_fat': ['importo', 'totale', 'amount', 'importo_fat'],
    'paese_fat': ['paese', 'country', 'nazione', 'paese_fat'],
    'mrn': ['mrn', 'numero mrn', 'mrn_id'],
    'fattura_dog': ['fattura', 'rif. fattura', 'invoice ref', 'fattura_dog'],
    'data_dog': ['data', 'data dichiarazione', 'data_dog'],
    'importo_dog': ['val. stat.', 'val.stat', 'valore stat', 'importo', 'valore', 'imponibile', 'importo_dog'],
    'cliente_dog': ['cliente', 'destinatario', 'consegnatario', 'cliente_dog'],
    'paese_dog': ['paese', 'country', 'paese origine', 'paese_dog']
  };

  // Prova a matchare le colonne
  const allHeaders = fatHeaders.concat(dogHeaders);
  const used = new Set();

  // Se "Val. Stat." è stato impostato, segnalo come usato per evitare doppi match
  if (valStatCol) used.add(valStatCol);

  Object.entries(simpleMap).forEach(([field, aliases]) => {
    // Salta importo_dog perché è già stato impostato
    if (field === 'importo_dog' && valStatCol) return;

    for (let h of allHeaders) {
      if (used.has(h)) continue;
      const lower = h.toLowerCase();
      if (aliases.some(alias => lower.includes(alias.toLowerCase()))) {
        const isdog = field.endsWith('_dog') || field === 'mrn';
        const map = isdog ? mapping.dog : mapping.fat;
        const cleanField = field.endsWith('_dog') ? field.replace('_dog', '') : field.replace('_fat', '');
        if (field === 'mrn') map[field] = h;
        else if (isdog) map[field] = h;
        else map[cleanField] = h;
        used.add(h);
        break;
      }
    }
  });

  // Applica le pre-selezioni ai select
  const panel = document.getElementById('colConfigPanel');
  panel.querySelectorAll('.col-select').forEach(sel => {
    const field = sel.dataset.field;
    const [type, fieldName] = field.split('_', 1)[0] === 'fat' ? ['fat', field.slice(4)] : ['dog', field.slice(4)];
    const map = type === 'fat' ? mapping.fat : mapping.dog;
    if (map[fieldName]) {
      sel.value = map[fieldName];
    }
  });
}

async function loadProfileList() {
  const sel = document.getElementById('selProfile');

  // Rimuovi tutte le option tranne la prima
  const options = sel.querySelectorAll('option');
  for (let i = options.length - 1; i > 0; i--) {
    options[i].remove();
  }

  if (!TOKEN) {
    console.warn('❌ Token non disponibile, nessun profilo caricato');
    return;
  }

  try {
    console.log('📥 Carico lista profili dal server...');
    // Carica profili dall'utente corrente (autenticato)
    const response = await fetch(`/api/profiles?token=${TOKEN}`);
    if (!response.ok) {
      console.warn('❌ Errore caricamento profili:', response.status);
      return;
    }
    const userProfiles = await response.json();
    console.log('✅ Profili ricevuti:', Object.keys(userProfiles));

    // SALVA I PROFILI IN CACHE (globale)
    cachedProfiles = userProfiles;
    console.log('💾 Profili salvati in cache:', Object.keys(cachedProfiles));

    // Popola dropdown con profili dell'utente
    Object.keys(userProfiles).forEach(name => {
      const opt = document.createElement('option');
      opt.value = name;
      opt.textContent = name;
      sel.appendChild(opt);
      console.log(`  ✓ Aggiunto profilo: ${name}`);
    });
  } catch (e) {
    console.error('❌ Errore fetch profili:', e);
  }
}

async function loadProfile(profileName = null) {
  const sel = document.getElementById('selProfile');
  const nameToLoad = profileName || sel.value;

  if (!nameToLoad) {
    console.log('❌ loadProfile: nameToLoad mancante');
    return;
  }

  try {
    console.log(`📥 Cerco profilo "${nameToLoad}" nella cache...`);

    // LEGGI DALLA CACHE (caricata da loadProfileList)
    const profile = cachedProfiles[nameToLoad];

    if (!profile) {
      console.warn(`❌ Profilo "${nameToLoad}" non trovato in cache. Cache disponibili:`, Object.keys(cachedProfiles));
      return;
    }

    console.log(`✅ Profilo "${nameToLoad}" trovato nella cache:`, profile);

    const panel = document.getElementById('colConfigPanel');
    const selectElements = panel.querySelectorAll('.col-select');
    selectElements.forEach(selectEl => {
      const field = selectEl.dataset.field; // es. 'fat_num_fattura' o 'dog_mrn'
      const [type, ...parts] = field.split('_');
      const fieldName = parts.join('_');
      const map = type === 'fat' ? profile.fat : profile.dog;
      if (map && map[fieldName]) {
        console.log(`  ✓ Imposto ${fieldName} = ${map[fieldName]}`);
        selectEl.value = map[fieldName];
      }
    });

    document.getElementById('inputProfileName').value = nameToLoad;

    // SALVA IL PROFILO COME ULTIMO USATO
    localStorage.setItem('mrn_last_profile', nameToLoad);

    // STAMPA IL PROFILO ALLA CONSOLE
    console.clear();
    console.log('%c═══════════════════════════════════════════════════════', 'color: #00c851; font-weight: bold;');
    console.log(`%cProfilo "${nameToLoad}" caricato:`, 'color: #00c851; font-weight: bold;');
    console.log('%c═══════════════════════════════════════════════════════', 'color: #00c851; font-weight: bold;');
    console.log('%cCopia questo JSON:', 'color: #ffbb33; font-weight: bold;');
    console.log(JSON.stringify(profile, null, 2));
    console.log('%c═══════════════════════════════════════════════════════', 'color: #00c851; font-weight: bold;');
  } catch (e) {
    console.error('❌ Errore caricamento profilo:', e);
  }
}

function saveProfile() {
  const name = document.getElementById('inputProfileName').value.trim();
  if (!name) {
    alert('Inserisci un nome per il profilo');
    return;
  }

  if (!TOKEN) {
    alert('Token non disponibile, impossibile salvare');
    return;
  }

  const mapping = getColMapping();

  // Salva il profilo nel SERVER (autenticato)
  fetch(`/api/profiles/${encodeURIComponent(name)}?token=${TOKEN}`, {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify(mapping)
  })
    .then(r => {
      if (!r.ok) throw new Error(r.statusText);
      return r.json();
    })
    .then(data => {
      loadProfileList();
      const sel = document.getElementById('selProfile');
      sel.value = name;
      localStorage.setItem('mrn_last_profile', name);
      alert('✓ Profilo salvato');
    })
    .catch(e => {
      console.error('Errore salvataggio profilo:', e);
      alert('✗ Errore salvataggio profilo');
    });
}

function deleteProfile(profileName) {
  if (!profileName || !TOKEN) {
    alert('Errore: profilo o token non disponibile');
    return;
  }

  if (!confirm(`Sei sicuro di voler eliminare il profilo "${profileName}"?`)) {
    return;
  }

  // Elimina il profilo dal SERVER
  fetch(`/api/profiles/${encodeURIComponent(profileName)}?token=${TOKEN}`, {
    method: 'DELETE'
  })
    .then(r => {
      if (!r.ok) throw new Error(r.statusText);
      return r.json();
    })
    .then(data => {
      alert('✓ Profilo eliminato');
      loadProfileList();
      document.getElementById('selProfile').value = '';
    })
    .catch(e => {
      console.error('Errore eliminazione profilo:', e);
      alert('✗ Errore eliminazione profilo');
    });
}

function getColMapping() {
  const mapping = { fat: {}, dog: {} };
  const panel = document.getElementById('colConfigPanel');
  panel.querySelectorAll('.col-select').forEach(sel => {
    if (!sel.value) return;
    const field = sel.dataset.field; // es. 'fat_num_fattura'
    const [type, ...parts] = field.split('_');
    const fieldName = parts.join('_');
    mapping[type][fieldName] = sel.value;
  });
  return mapping;
}

// ── RICONCILIAZIONE ────────────────────────────────────────────────────────
async function doReconcile() {
  const btn = document.getElementById('btnRec');
  const spin = document.getElementById('spinRec');
  const alert = document.getElementById('alertRec');
  btn.disabled = true; spin.style.display = ''; alert.innerHTML = '';

  try {
    const fd = new FormData();
    fd.append('file_dogana', fileDogana);
    fd.append('file_fatture', fileFatture);

    // Aggiungi mapping se configurato
    const mapping = getColMapping();
    // Salva il mapping corrente per l'export
    currentColMapping = mapping;
    if (Object.keys(mapping.fat).length > 0 || Object.keys(mapping.dog).length > 0) {
      fd.append('col_mapping', JSON.stringify(mapping));
    }

    const r = await fetch('/api/reconcile', {
      method: 'POST', body: fd, headers: { 'X-Token': TOKEN }
    });
    if (!r.ok) {
      const err = await r.json();
      throw new Error(err.detail || 'Errore server');
    }
    const data = await r.json();
    lastResults = data;
    renderStats(data.stats);
    renderResults(data.results);
    document.getElementById('btnExport').disabled = false;

    // Aggiorna quota
    fetch('/api/me', { headers: { 'X-Token': TOKEN } }).then(r=>r.json()).then(updateQuota);
  } catch(e) {
    alert.innerHTML = `<div class="alert alert-error">✗ ${e.message}</div>`;
  } finally {
    btn.disabled = false; spin.style.display = 'none';
    checkReady();
  }
}

function renderStats(s) {
  document.getElementById('statsArea').style.display = '';
  document.getElementById('s-tot').textContent  = s.n_fatture;
  document.getElementById('s-mrn').textContent  = s.n_mrn;
  document.getElementById('s-cert').textContent = s.CERTO;
  document.getElementById('s-prob').textContent = s.PROBABILE;
  document.getElementById('s-poss').textContent = s.POSSIBILE;
  document.getElementById('s-deb').textContent  = s.DEBOLE;
  document.getElementById('s-none').textContent = s.NESSUN_MATCH;
}

function renderResults(results) {
  const area = document.getElementById('resultsArea');
  area.innerHTML = '';
  results.forEach((fat, fi) => {
    const topCand = fat.candidati[0];
    const topClass = topCand ? topCand.class : 'NESSUN MATCH';
    const topScore = topCand ? topCand.score : 0;

    const block = document.createElement('div');
    block.className = 'fat-block';

    const hdr = document.createElement('div');
    hdr.className = 'fat-header';
    hdr.innerHTML = `
      <div class="fat-info">
        <div class="fat-num">📄 ${fat.fattura_num || '(n/d)'}</div>
        <div class="fat-meta">
          ${fat.fattura_data ? '📅 ' + fat.fattura_data : ''}
          ${fat.fattura_cliente ? ' · 🏢 ' + fat.fattura_cliente : ''}
          ${fat.fattura_importo ? ' · 💶 ' + fat.fattura_importo : ''}
          · ${fat.candidati.length} candidat${fat.candidati.length === 1 ? 'o' : 'i'}
        </div>
      </div>
      <span class="fat-badge badge-${topClass}">${topClass}</span>
      <span style="color:var(--muted);font-size:.85rem;min-width:40px;text-align:right">${topScore}</span>
      <span style="color:var(--muted)">▾</span>
    `;

    const cands = document.createElement('div');
    cands.className = 'fat-candidates';

    fat.candidati.forEach((c, ci) => {
      const row = document.createElement('div');
      row.className = 'candidate-row';
      row.innerHTML = `
        <div class="cand-rank">#${ci+1}</div>
        <div class="cand-mrn">${c.mrn}</div>
        <div class="cand-score" style="color:var(--${c.class==='CERTO'?'certo':c.class==='PROBABILE'?'prob':c.class==='POSSIBILE'?'poss':c.class==='DEBOLE'?'debole':'none'})">${c.score}</div>
        <div class="cand-class"><span class="fat-badge badge-${c.class}">${c.class}</span></div>
        <div class="cand-notes">
          ${c.explanation}
          <div class="cand-method">${c.method}</div>
        </div>
      `;
      cands.appendChild(row);
    });

    if (fat.candidati.length === 0) {
      cands.innerHTML = '<div style="padding:12px 16px;color:var(--muted);font-size:.85rem">Nessun MRN candidato con score ≥ 50</div>';
    }

    hdr.onclick = () => {
      cands.classList.toggle('open');
      hdr.querySelector('span:last-child').textContent =
        cands.classList.contains('open') ? '▴' : '▾';
    };

    // Apri automaticamente se pochi risultati o score alto
    if (results.length <= 10 || topScore >= 85) cands.classList.add('open');

    block.appendChild(hdr);
    block.appendChild(cands);
    area.appendChild(block);
  });
}

// ── EXPORT ─────────────────────────────────────────────────────────────────
async function doExport() {
  const btn = document.getElementById('btnExport');
  const spin = document.getElementById('spinExp');
  btn.disabled = true; spin.style.display = '';
  try {
    const fd = new FormData();
    fd.append('file_dogana', fileDogana);
    fd.append('file_fatture', fileFatture);

    // Usa il mapping salvato durante la riconciliazione
    if (currentColMapping && (Object.keys(currentColMapping.fat).length > 0 || Object.keys(currentColMapping.dog).length > 0)) {
      fd.append('col_mapping', JSON.stringify(currentColMapping));
    }

    const r = await fetch('/api/export', {
      method: 'POST', body: fd, headers: { 'X-Token': TOKEN }
    });
    if (!r.ok) throw new Error((await r.json()).detail);
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = r.headers.get('Content-Disposition')?.split('filename=')[1] || 'Riconciliazione_MRN.xlsx';
    a.click();
    URL.revokeObjectURL(url);
    fetch('/api/me', { headers: { 'X-Token': TOKEN } }).then(r=>r.json()).then(updateQuota);
  } catch(e) {
    document.getElementById('alertRec').innerHTML = `<div class="alert alert-error">Export: ${e.message}</div>`;
  } finally {
    btn.disabled = false; spin.style.display = 'none';
  }
}

// ── ADMIN: TOKEN ───────────────────────────────────────────────────────────
function copyTokenToClipboard(tok) {
  navigator.clipboard.writeText(tok).then(() => {
    alert('✅ Token copiato negli appunti!');
  }).catch(err => {
    alert('❌ Errore nella copia: ' + err.message);
  });
}

async function loadTokens() {
  const r = await fetch('/api/admin/tokens', { headers: { 'X-Token': TOKEN } });
  const tokens = await r.json();
  const tbody = document.getElementById('tokenTbody');
  tbody.innerHTML = '';
  tokens.forEach(t => {
    const tr = document.createElement('tr');
    const pct = t.max_rows > 0 ? Math.round((t.rows_used / t.max_rows) * 100) : 0;
    tr.innerHTML = `
      <td class="tok-cell" title="Clicca per copiare" style="cursor:pointer;font-family:monospace;font-size:.85rem;white-space:nowrap" onclick="copyTokenToClipboard('${t.token}')">${t.token.slice(0,12)}…</td>
      <td style="white-space:nowrap">${t.name}</td>
      <td style="white-space:nowrap">${t.rows_used}${t.max_rows > 0 ? '/' + t.max_rows : ''}</td>
      <td style="white-space:nowrap">${t.remaining}</td>
      <td style="white-space:nowrap">
        <span style="color:${t.active?'var(--certo)':'var(--debole)'}">
          ${t.active ? '● Attivo' : '○ Sospeso'}
        </span>
      </td>
      <td style="color:var(--muted);font-size:.78rem;white-space:nowrap">${t.created_at ? t.created_at.slice(0,10) : ''}</td>
      <td style="color:var(--muted);font-size:.78rem">${t.notes || ''}</td>
      <td style="white-space:nowrap">
        <button class="btn btn-ghost" style="font-size:.75rem;padding:4px 6px"
          onclick="showTokenModal('${t.token}', '${t.name}')">
          📋
        </button>
        <button class="btn btn-ghost" style="font-size:.75rem;padding:4px 6px"
          onclick="editTokenValue('${t.token}')">
          ✏️
        </button>
        <button class="btn btn-ghost" style="font-size:.75rem;padding:4px 6px"
          onclick="toggleToken('${t.token}', ${!t.active})">
          ${t.active ? '⏸' : '▶'}
        </button>
        <button class="btn btn-ghost" style="font-size:.75rem;padding:4px 6px"
          onclick="extendToken('${t.token}', ${t.max_rows})">
          +
        </button>
      </td>
    `;
    tbody.appendChild(tr);
  });
}

async function createToken() {
  const name    = document.getElementById('newName').value.trim();
  const maxRows = parseInt(document.getElementById('newMaxRows').value) || 1000;
  const notes   = document.getElementById('newNotes').value.trim();
  if (!name) { alert('Inserisci un nome'); return; }
  const r = await fetch('/api/admin/tokens', {
    method: 'POST',
    headers: { 'X-Token': TOKEN, 'Content-Type': 'application/json' },
    body: JSON.stringify({ name, max_rows: maxRows, notes })
  });
  const data = await r.json();
  const div = document.getElementById('newTokenResult');
  div.innerHTML = `<div class="alert alert-success">Token generato — Copia e invia al cliente:<br>
    <strong style="cursor:pointer" onclick="navigator.clipboard.writeText('${data.token}')">${data.token}</strong>
    <br><small style="color:var(--muted)">Clicca per copiare</small></div>`;
  loadTokens();
}

async function toggleToken(tok, newActive) {
  await fetch(`/api/admin/tokens/${tok}`, {
    method: 'PATCH',
    headers: { 'X-Token': TOKEN, 'Content-Type': 'application/json' },
    body: JSON.stringify({ active: newActive })
  });
  loadTokens();
}

async function extendToken(tok, currentMax) {
  const add = parseInt(prompt('Quante righe aggiungere?', '500'));
  if (!add || add <= 0) return;
  await fetch(`/api/admin/tokens/${tok}`, {
    method: 'PATCH',
    headers: { 'X-Token': TOKEN, 'Content-Type': 'application/json' },
    body: JSON.stringify({ max_rows: currentMax + add })
  });
  loadTokens();
}

function showTokenModal(tok, name) {
  const modal = document.createElement('div');
  modal.style.cssText = 'position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.6);display:flex;align-items:center;justify-content:center;z-index:9999';
  modal.innerHTML = `
    <div style="background:var(--bg);padding:24px;border-radius:8px;max-width:500px;box-shadow:0 8px 32px rgba(0,0,0,0.3);border:1px solid var(--border)">
      <h3 style="margin-top:0">Token: ${name}</h3>
      <div style="background:var(--bg-alt);padding:12px;border-radius:4px;word-break:break-all;font-family:monospace;font-size:.9rem;margin:16px 0;color:var(--accent2)">
        ${tok}
      </div>
      <button class="btn btn-primary" style="width:100%;margin-bottom:8px" onclick="navigator.clipboard.writeText('${tok}'); alert('Token copiato negli appunti!')">📋 Copia Token</button>
      <button class="btn btn-ghost" style="width:100%" onclick="this.parentElement.parentElement.remove()">Chiudi</button>
    </div>
  `;
  document.body.appendChild(modal);
  modal.onclick = (e) => { if (e.target === modal) modal.remove(); };
}

async function editTokenValue(oldTok) {
  const newTok = prompt('Inserisci il nuovo valore del token:', oldTok);
  if (!newTok || newTok === oldTok || newTok.trim() === '') return;
  if (!confirm('Sei sicuro di voler cambiare il token? I vecchi token non funzioneranno più.')) return;

  try {
    const r = await fetch(`/api/admin/tokens/${oldTok}`, {
      method: 'PATCH',
      headers: { 'X-Token': TOKEN, 'Content-Type': 'application/json' },
      body: JSON.stringify({ token: newTok.trim() })
    });
    if (r.ok) {
      alert('Token aggiornato con successo!');
      loadTokens();
    } else {
      const err = await r.json();
      alert('Errore: ' + (err.detail || 'Impossibile aggiornare il token'));
    }
  } catch(e) {
    alert('Errore: ' + e.message);
  }
}

// ── ADMIN: LOGS ────────────────────────────────────────────────────────────
async function loadLogs() {
  const r = await fetch('/api/admin/logs?limit=100', { headers: { 'X-Token': TOKEN } });
  const logs = await r.json();
  const area = document.getElementById('logArea');
  if (!logs.length) { area.innerHTML = '<div style="color:var(--muted);font-size:.85rem">Nessuna attività registrata</div>'; return; }
  area.innerHTML = logs.map(l => {
    let detail = '';
    if (l.action === 'reconcile') {
      detail = `${l.n_fatture}f × ${l.n_mrn}MRN → C:${l.stats?.CERTO} P:${l.stats?.PROBABILE} PS:${l.stats?.POSSIBILE} D:${l.stats?.DEBOLE} N:${l.stats?.NO_MATCH}`;
    } else if (l.action === 'export_excel') {
      detail = l.filename;
    } else if (l.action === 'create_token') {
      detail = `${l.new_token_name} (${l.max_rows} righe)`;
    } else {
      detail = JSON.stringify(l).slice(0, 80);
    }
    return `<div class="log-entry">
      <span class="log-ts">${l.ts}</span>
      <span class="log-user">${l.user}</span>
      <span class="log-action">${l.action}</span>
      <span class="log-detail">${detail}</span>
    </div>`;
  }).join('');
}

// ════════════════════════════════════════════════════════════════════════════
// ── IVISTO CC599C VIEWER ────────────────────────────────────────────────────
// ════════════════════════════════════════════════════════════════════════════

const EXIT_CODES = {
  A1:{ label:"Conforme (A1)",              desc:"Merce uscita dal territorio doganale UE a seguito di controllo fisico (totale o parziale) — nessuna irregolarita' rilevata.",                                                                                                                                      cls:"A1", dot:"info",    badge:"badge-info" },
  A2:{ label:"Ritenuto conforme (A2)",     desc:"Merce uscita dal territorio doganale UE a seguito di controllo esclusivamente documentale, senza ispezione fisica, oppure senza alcun controllo — nessuna irregolarita'. Costituisce prova di esportazione valida ai fini IVA (art. 8 DPR 633/72).",           cls:"A2", dot:"ok",      badge:"badge-ok" },
  A3:{ label:"Procedura semplificata (A3)",desc:"Uscita confermata tramite procedura semplificata — nessun controllo eseguito dall'Ufficio di Uscita (spedizioniere autorizzato / AEO).",                                                                                                                      cls:"A3", dot:"warn",    badge:"badge-warn" },
  A4:{ label:"Difformita' minori (A4)",    desc:"Rilevate difformita' minori durante il controllo. La merce e' rilasciata, ma l'Ufficio Doganale di Supervisione (SCO) verifica le discrepanze e puo' informare il dichiarante al di fuori del sistema. La dichiarazione non puo' essere emendata.",          cls:"A3", dot:"warn",    badge:"badge-warn" },
  A5:{ label:"Sigilli non conformi (A5)",  desc:"I sigilli non sono risultati conformi (assenti, danneggiati o con discrepanze). La merce puo' essere comunque rilasciata se le difformita' sono minori.",                                                                                                     cls:"A3", dot:"warn",    badge:"badge-warn" },
  B1:{ label:"Difformita' maggiori (B1)",  desc:"Rilevate difformita' maggiori durante il controllo doganale all'uscita. Contattare l'Ufficio Doganale di Esportazione.",                                                                                                                                      cls:"B",  dot:"danger",  badge:"badge-err" },
  B2:{ label:"Non trovato (B2)",           desc:"La merce non e' stata trovata al momento dell'uscita dal territorio doganale UE.",                                                                                                                                                                             cls:"B",  dot:"danger",  badge:"badge-err" },
  B3:{ label:"Non applicabile (B3)",       desc:"Risultato non applicabile per la tipologia di operazione.",                                                                                                                                                                                                   cls:"default", dot:"neutral", badge:"badge-neutral" },
};

const NATIONS = {
  IT:"Italia",AT:"Austria",BE:"Belgio",CH:"Svizzera",CZ:"Rep. Ceca",DE:"Germania",
  DK:"Danimarca",ES:"Spagna",FI:"Finlandia",FR:"Francia",GB:"Regno Unito",GR:"Grecia",
  HR:"Croazia",HU:"Ungheria",IE:"Irlanda",LU:"Lussemburgo",NL:"Paesi Bassi",NO:"Norvegia",
  PL:"Polonia",PT:"Portogallo",RO:"Romania",SE:"Svezia",SI:"Slovenia",SK:"Slovacchia",
  TR:"Turchia",US:"USA",CN:"Cina",JP:"Giappone",MA:"Marocco",EG:"Egitto",
};

const OFFICES = {
  IT273000:"UADM Lombardia 4 — Varese (ufficio principale)",IT273100:"UADM Lombardia 4 — Varese, sezione centrale (Via Dalmazia 56)",IT273102:"UADM Lombardia 4 — Distaccamento locale Gaggiolo (Cantello VA)",IT273104:"UADM Lombardia 4 — Sezione operativa Gaggiolo",IT273105:"UADM Lombardia 4 — Sezione operativa Luino",IT273199:"UADM Lombardia 4 — Varese, cassa centrale",IT232100:"UADM Lombardia 4 — Varese, sezione operativa (codice da verificare)",IT275000:"UADM Lombardia 5 — Como (ufficio principale)",IT275100:"UADM Lombardia 5 — Como, sezione centrale",IT275102:"UADM Lombardia 5 — Sezione operativa Lecco",IT275103:"UADM Lombardia 5 — Sezione operativa Montano Lucino",IT275104:"UADM Lombardia 5 — Sezione operativa Oria Valsolda",IT275105:"UADM Lombardia 5 — Sezione operativa Ponte Chiasso",IT275199:"UADM Lombardia 5 — Como, cassa centrale",IT277000:"UADM Lombardia 1 — Milano 1 (ufficio principale)",IT277100:"UADM Lombardia 1 — Milano 1, sezione centrale",IT277199:"UADM Lombardia 1 — Milano 1, cassa centrale",IT278000:"UADM Lombardia 2 — Milano 2 (ufficio principale)",IT278100:"UADM Lombardia 2 — Milano 2, sezione centrale",IT278199:"UADM Lombardia 2 — Milano 2, cassa centrale",IT279000:"UADM Lombardia 3 — Malpensa (ufficio principale)",IT279100:"UADM Lombardia 3 — Malpensa, sezione aeroportuale",IT279199:"UADM Lombardia 3 — Malpensa, cassa centrale",IT371000:"UADM Lombardia 6 — Milano 3 (ufficio principale)",IT371100:"UADM Lombardia 6 — Milano 3, sezione centrale",IT371199:"UADM Lombardia 6 — Milano 3, cassa centrale",IT271000:"UADM Lombardia — Pavia",IT272000:"UADM Lombardia — Tirano",IT274000:"UADM Lombardia — Brescia",IT274100:"UADM Lombardia — Brescia, sezione centrale",IT274101:"UADM Lombardia — Sezione operativa Cremona",IT276000:"UADM Lombardia — Bergamo",IT276100:"UADM Lombardia — Bergamo, sezione centrale",IT079000:"UADM — Mantova",IT079100:"UADM — Mantova, sezione centrale",IT223000:"UADM Emilia 1 — Bologna (sede)",IT223100:"UADM Emilia 1 Bologna — area centrale",IT223101:"UADM Emilia 1 Bologna — Aeroporto G. Marconi",IT223102:"UADM Emilia 1 Bologna — Interporto",IT223103:"UADM Emilia 1 Bologna — area territoriale Ferrara",IT029000:"UADM Emilia 2 — Piacenza (sede)",IT029100:"UADM Emilia 2 Piacenza — area centrale",IT028000:"UADM Emilia 3 — Parma (sede)",IT028100:"UADM Emilia 3 Parma — area centrale",IT028101:"UADM Emilia 3 Parma — Aeroporto G. Verdi",IT224000:"UADM Emilia 4 — Modena (sede)",IT224100:"UADM Emilia 4 Modena — area centrale",IT224101:"UADM Emilia 4 Modena — area territoriale Reggio Emilia",IT221000:"UADM Romagna 1 — Ravenna (sede)",IT221100:"UADM Romagna 1 Ravenna — area centrale",IT225000:"UADM Romagna 2 — Rimini (sede)",IT225100:"UADM Romagna 2 Rimini — area centrale",IT225101:"UADM Romagna 2 Rimini — area territoriale Forli-Cesena",IT222000:"Bologna — cod. precedente (-> IT223000 dal 04/2025)",IT025000:"Modena — cod. precedente (-> IT224000 dal 04/2025)",IT026000:"Rimini — cod. precedente (-> IT225000 dal 04/2025)",IT022999:"Reggio Emilia — cod. precedente (-> IT224101 dal 04/2025)",IT024000:"Forli-Cesena — cod. precedente (-> IT225101 dal 04/2025)",IT027000:"Ferrara — cod. precedente (-> IT223103 dal 04/2025)",IT118000:"Aosta",IT116000:"Biella",IT311000:"Cuneo",IT312000:"Novara",IT313000:"Alessandria",IT314000:"Torino",IT117000:"Vercelli",IT119000:"Verbano-Cusio-Ossola",IT261000:"Genova 1",IT262000:"Genova 2",IT263000:"Rivalta Scrivia — Retroporto Genova",IT067000:"Imperia",IT068000:"La Spezia",IT066000:"Savona",IT034000:"Bolzano",IT035000:"Trento",IT134000:"Treviso",IT135000:"Vicenza",IT136000:"Verona",IT137000:"Venezia",IT138000:"Padova",IT126000:"Pordenone",IT127000:"Gorizia",IT128000:"Udine",IT129000:"Trieste",IT321000:"Fernetti — Retroporto Trieste",IT051999:"Arezzo",IT057000:"Firenze",IT055000:"Livorno",IT054000:"Pisa",IT056000:"Prato e Pistoia",IT107000:"Perugia",IT108000:"Terni",IT305000:"Ancona",IT308000:"Civitanova Marche",IT301000:"Civitavecchia",IT302000:"Frosinone",IT105999:"Gaeta",IT309000:"L'Aquila",IT303000:"Campobasso",IT304000:"Pescara",IT306000:"Roma 1",IT307000:"Roma 2",IT088000:"Caserta",IT089000:"Benevento",IT281000:"Napoli 1",IT282000:"Napoli 2",IT084000:"Salerno",IT019000:"Potenza",IT014000:"Foggia",IT015000:"Brindisi",IT016000:"Lecce",IT017000:"Taranto",IT018000:"Bari",IT085000:"Catanzaro",IT086000:"Gioia Tauro",IT087000:"Reggio Calabria",IT291000:"Catania",IT292000:"Palermo",IT096000:"Porto Empedocle",IT097000:"Trapani",IT098000:"Messina",IT099000:"Siracusa",IT043000:"Sassari",IT044000:"Cagliari",IT922103:"ADM — Ufficio AEO e semplificazioni",IT922104:"ADM — Ufficio Tariffa e Dazi",IT922105:"ADM — Ufficio Tributi Doganali",IT922106:"ADM — Ufficio Regimi e Traffici di Confine",CH001731:"Zoll Nord — Pratteln / Basel (CH)",CH001801:"Zoll Nord — Basel/Weil Rhein-Autobahn (CH)",CH001841:"Zoll Nord — Basel/St. Louis Autobahn (CH)",CH003110:"Basel/Mulhouse Airport (CH/FR)",CH004031:"Dogana Ticino — Chiasso Strada (CH)",CH004041:"Dogana Ticino — Chiasso Brogeda (CH)",CH004061:"Dogana Ticino — Stabio/Gaggiolo (CH)",CH004071:"Dogana Ticino — Ponte Tresa (CH)",CH004081:"Dogana Ticino — Zenna/Luino (CH)",CH006031:"Dogana Est — St. Margrethen (CH)",CH007011:"Dogana Est — Buchs SG (CH)",NL000396:"Amsterdam Airport Schiphol (NL)",NL000399:"Rotterdam — Maasvlakte (NL)",NL003000:"Rotterdam — Botlek (NL)",DE004700:"Hamburg Airport (DE)",DE006200:"Frankfurt Airport (DE)",DE008000:"Muenchen Airport (DE)",FR001300:"Paris CDG Airport (FR)",FR002100:"Marseille Port (FR)",BE000100:"Bruxelles Nationaal Airport (BE)",BE000200:"Antwerpen Port (BE)",ES000100:"Barcelona Port (ES)",ES001300:"Madrid Barajas Airport (ES)",GB000060:"Heathrow Airport (GB)",GB003110:"Dover Port (GB)",PL000100:"Warszawa Okecie Airport (PL)",AT000100:"Wien Flughafen (AT)",SE000100:"Stockholm Arlanda (SE)",DK000200:"Copenhagen Airport (DK)",FI000100:"Helsinki-Vantaa Airport (FI)",
};

function getOffice(code) {
  if (!code) return { name:"—", code:"", unknown:false };
  const name = OFFICES[code];
  if (name) return { name, code, unknown:false };
  const nation = NATIONS[code.substring(0,2)];
  return { name: nation ? "Ufficio doganale — " + nation : code, code, unknown:true };
}

function ivistoFmtDate(d) {
  if (!d) return "—";
  try {
    const p = d.includes("T") ? d.split("T")[0] : d;
    const [y,m,dd] = p.split("-");
    const mo = ["gen","feb","mar","apr","mag","giu","lug","ago","set","ott","nov","dic"];
    return parseInt(dd) + " " + mo[parseInt(m)-1] + " " + y;
  } catch(e) { return d; }
}

function ivSb(label, value, full, mono) {
  return '<div class="stat-box' + (full?" full":"") + '">' +
    '<div class="stat-label">' + label + '</div>' +
    '<div class="stat-value' + (mono?" mono":"") + '">' + value + '</div>' +
    '</div>';
}

function officeBox(label, code, full) {
  const o = getOffice(code);
  const unk = o.unknown ? '<span class="tag-unknown">non in rubrica</span>' : "";
  const sub = o.code ? '<div class="stat-sub">' + o.code + '</div>' : "";
  return '<div class="stat-box' + (full?" full":"") + '">' +
    '<div class="stat-label">' + label + '</div>' +
    '<div class="stat-value">' + o.name + unk + '</div>' +
    sub + '</div>';
}

function renderDoc(xml) {
  let doc;
  try { doc = new DOMParser().parseFromString(xml, "application/xml"); } catch(e) { return null; }
  if (doc.getElementsByTagName("parsererror").length) return null;

  const req = {
    "MRN":"MRN (Movement Reference Number)",
    "ExitControlResult":"ExitControlResult (esito controllo uscita)",
    "preparationDateAndTime":"preparationDateAndTime (data conclusione uscita)",
    "referenceNumber":"referenceNumber (codice ufficio doganale)",
  };
  const missing = Object.entries(req).filter(([t]) => !doc.getElementsByTagName(t).length).map(([,l]) => l);
  if (missing.length) return { error:"Il file XML caricato non e' un messaggio CC599C (IVISTO).<br>Tag obbligatori mancanti:<ul>" + missing.map(m => "<li>" + m + "</li>").join("") + "</ul>" };

  const refs = [...doc.getElementsByTagName("referenceNumber")];
  const exportRef = refs[0]?.textContent || null;
  const exitRef = refs[1]?.textContent || null;
  const destNode = doc.getElementsByTagName("CustomsOfficeOfDestination")[0];
  const destRef = destNode ? destNode.getElementsByTagName("referenceNumber")[0]?.textContent : null;
  const mrn = doc.getElementsByTagName("MRN")[0]?.textContent || null;
  const transit = doc.getElementsByTagName("transit")[0]?.textContent === "1";
  const rc = doc.getElementsByTagName("code")[0]?.textContent || null;
  const exitDate = doc.getElementsByTagName("exitDate")[0]?.textContent || null;
  const prepDate = doc.getElementsByTagName("preparationDateAndTime")[0]?.textContent || null;
  const msgId = doc.getElementsByTagName("messageIdentification")[0]?.textContent || null;
  const sender = doc.getElementsByTagName("messageSender")[0]?.textContent || null;
  const suppDocs = [...doc.getElementsByTagName("SupportingDocument")].map(sd => ({type:sd.getElementsByTagName("type")[0]?.textContent || "", ref:sd.getElementsByTagName("referenceNumber")[0]?.textContent || ""}));

  const code = EXIT_CODES[rc] || { label:"Codice "+(rc||"?"), desc:"Codice esito non riconosciuto (tabella CL393).", cls:"default", dot:"neutral", badge:"badge-neutral" };
  const rcls = rc ? "result-"+rc.charAt(0) : "result-default";

  const suppHTML = suppDocs.length ? '<div class="doc-section">' +
    '<div class="doc-section-title">Documenti di supporto (' + suppDocs.length + ')</div>' +
    suppDocs.map(sd => '<div class="stat-grid" style="margin-bottom:8px">' + ivSb("Tipo documento", sd.type) + ivSb("Riferimento", sd.ref, false, true) + '</div>').join("") + '</div>' : "";

  return '<div class="card">' +
    '<div class="card-head"><span>Notifica di Uscita &mdash; CC599C</span>' +
      '<span class="badge ' + code.badge + '"><span class="dot dot-' + code.dot + '" style="margin-right:4px"></span>' + code.label + '</span>' +
    '</div>' +
    '<div class="doc-section">' +
      '<div class="doc-section-title">Esito controllo uscita</div>' +
      '<div class="result-block ' + rcls + '">' +
        '<div class="result-code">' + (rc||"—") + ' &mdash; ' + code.label + '</div>' +
        '<div class="result-desc">' + code.desc + '</div>' +
      '</div>' +
      '<div class="stat-grid">' +
        ivSb("Data uscita effettiva", ivistoFmtDate(exitDate)) +
        ivSb("Uscita conclusa in data", ivistoFmtDate(prepDate)) +
        ivSb("Movimento in transito", transit ? "Si" : "No") +
      '</div>' +
    '</div>' +
    '<div class="doc-section">' +
      '<div class="doc-section-title">Dichiarazione di esportazione</div>' +
      '<div class="stat-grid">' +
        ivSb("MRN &mdash; Movement Reference Number", mrn||"—", true, true) +
        officeBox("Ufficio doganale di esportazione", exportRef) +
        officeBox("Ufficio doganale di uscita effettiva", exitRef) +
        (destRef ? officeBox("Ufficio doganale di destinazione (transito)", destRef, true) : "") +
      '</div>' +
    '</div>' +
    suppHTML +
    '<div class="doc-section">' +
      '<div class="doc-section-title">Dati messaggio</div>' +
      '<div class="stat-grid">' +
        ivSb("Tipo messaggio", "CC599C") +
        ivSb("Mittente sistema", sender||"—") +
        ivSb("ID messaggio", msgId||"—", true, true) +
      '</div>' +
    '</div>' +
    '<div style="border-top:1px solid var(--border);padding-top:12px;margin-top:4px">' +
      '<p class="note-text">Il messaggio CC599C e\' la notifica ufficiale di uscita delle merci dal territorio doganale dell\'Unione Europea, generata dal sistema AES/ECS2 PLUS. ' +
        'Sostituisce il precedente IE599 e costituisce prova di esportazione valida ai fini dell\'aliquota IVA zero ex art. 8 DPR 633/72 (per esiti A1 e A2). ' +
        'Codici uffici aggiornati ad aprile 2025 &mdash; fonti: ADM lista uffici ufficiale + avviso CNSD 02/04/2025.</p>' +
      '<p id="printDate" style="display:none;font-size:10px;color:var(--text3);margin-top:6px;font-family:var(--mono)">Stampato il: <span id="printDateVal"></span></p>' +
    '</div>' +
  '</div>';
}

function ivistoClearDoc() {
  ["out","errBox"].forEach(id => { const el=document.getElementById(id); el.style.display="none"; el.innerHTML=""; });
  document.getElementById("tb").style.display = "none";
  document.getElementById("uploadCard").style.display = "block";
  document.getElementById("fi").value = "";
  document.getElementById("xmlPaste").value = "";
  ivistoSwitchTab("file");
}

function processXML(text) {
  const result = renderDoc(text);
  if (!result) {
    document.getElementById("errBox").innerHTML = '<div class="alert alert-err">File XML non valido o non leggibile.</div>';
    document.getElementById("errBox").style.display = "block";
    return;
  }
  if (result.error) {
    document.getElementById("errBox").innerHTML = '<div class="alert alert-err">' + result.error + '</div>';
    document.getElementById("errBox").style.display = "block";
    document.getElementById("uploadCard").style.display = "block";
    return;
  }
  document.getElementById("out").innerHTML = result;
  document.getElementById("out").style.display = "block";
  document.getElementById("tb").style.display = "flex";
  document.getElementById("uploadCard").style.display = "none";
  document.getElementById("errBox").style.display = "none";
}

function ivistoSwitchTab(name) {
  document.querySelectorAll("#ivisto-panel .tab-btn").forEach((b,i) => b.classList.toggle("active", ["file","paste"][i] === name));
  document.getElementById("tab-file").classList.toggle("active", name === "file");
  document.getElementById("tab-paste").classList.toggle("active", name === "paste");
}

function ivistoAnalyzeText() {
  const text = document.getElementById("xmlPaste").value.trim();
  if (!text) return;
  processXML(text);
}

document.getElementById("fi").addEventListener("change", function() {
  if (!this.files[0]) return;
  const r = new FileReader(); r.onload = e => processXML(e.target.result); r.readAsText(this.files[0]);
});

const dz = document.getElementById("dropZone");
dz.addEventListener("dragover",  e => { e.preventDefault(); dz.style.borderColor = "var(--accent)"; });
dz.addEventListener("dragleave", () => { dz.style.borderColor = "var(--border2)"; });
dz.addEventListener("drop", e => {
  e.preventDefault(); dz.style.borderColor = "var(--border2)";
  const f = e.dataTransfer.files[0]; if (!f) return;
  const r = new FileReader(); r.onload = ev => processXML(ev.target.result); r.readAsText(f);
});

window.addEventListener("beforeprint", () => {
  const el = document.getElementById("printDate"), val = document.getElementById("printDateVal");
  if (!el||!val) return;
  const now = new Date(), mo = ["gen","feb","mar","apr","mag","giu","lug","ago","set","ott","nov","dic"];
  val.textContent = now.getDate() + " " + mo[now.getMonth()] + " " + now.getFullYear();
  el.style.display = "block";
});
window.addEventListener("afterprint", () => {
  const el = document.getElementById("printDate"); if (el) el.style.display = "none";
});
</script>
</body>
</html>
"""

@app.get("/", response_class=HTMLResponse)
async def home():
    return HTML


# ════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    log.info("=" * 60)
    log.info("  SIRIO - Sistema di Riconciliazione — avvio su http://0.0.0.0:8002")
    log.info(f"  Data dir: {DATA_DIR}")
    log.info(f"  Token file: {TOKENS_FILE}")
    log.info("=" * 60)
    uvicorn.run("main:app", host="0.0.0.0", port=8002, reload=False)
